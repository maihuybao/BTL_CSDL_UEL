import sys
import os
import sqlite3
from PyQt6 import QtWidgets, QtGui, QtCore
from PyQt6.QtWidgets import QApplication, QMainWindow, QDialog, QWidget, QTableWidgetItem, QMessageBox

# Tự động cấu hình sys.path để tránh lỗi import "unknown location" trên macOS/Windows
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
for path in [current_dir, parent_dir]:
    if path not in sys.path:
        sys.path.insert(0, path)

from Login.Login import Ui_LoginDialog

try:
    from Dashboard.Dashboard import Ui_MainWindow
except ModuleNotFoundError:
    from Dashboard.Dashboard import Ui_MainWindow

from Comment.Comment import Ui_CommentWidget
from Customer.Customer import Ui_CustomerWidget
from Livestream.Livestream import Ui_LivestreamWidget
from LivestreamDetail.LivestreamDetail import Ui_LivestreamDetailWidget
from Order.Order import Ui_OrderWidget
from OrderDetail.OrderDetail import Ui_OrderDetailWidget
from Payment.Payment import Ui_PaymentWidget
from Product.Product import Ui_ProductWidget
from Seller.Seller import Ui_SellerWidget
from Statistics.Statistics import Ui_StatisticsWidget
from Voucher.Voucher import Ui_VoucherWidget

DB_PATH = os.path.join(parent_dir, "data_db.sqlite")
if not os.path.exists(DB_PATH):
    DB_PATH = os.path.join(current_dir, "data_db.sqlite")

# QtCharts là tùy chọn: thiếu thì app vẫn chạy, chỉ không vẽ biểu đồ
try:
    from PyQt6.QtCharts import (QChart, QChartView, QBarSeries, QBarSet,
                                QBarCategoryAxis, QValueAxis, QPieSeries)
    HAS_QTCHARTS = True
except ImportError:
    HAS_QTCHARTS = False

# qtawesome (Font Awesome): thay toàn bộ emoji trên UI bằng icon chuẩn
try:
    import qtawesome as qta
    HAS_QTA = True
except ImportError:
    HAS_QTA = False

import re as _re

# Bộ icon Font Awesome cho các nút chức năng (khóa = tên objectName của nút)
_BUTTON_ICONS = {
    "btnThem": ("fa5s.plus", "Thêm mới"),
    "btnSua": ("fa5s.edit", "Sửa"),
    "btnXoa": ("fa5s.trash-alt", "Xóa bỏ"),
    "btnLuu": ("fa5s.save", "Lưu lại"),
    "btnHuy": ("fa5s.times", "Hủy bỏ"),
    "btnRefresh": ("fa5s.sync", "Làm mới"),
    "btnSearch": ("fa5s.search", "Tìm kiếm"),
    "btnXuatExcel": ("fa5s.file-excel", "Xuất Excel"),
    "btnFilter": ("fa5s.filter", "Lọc dữ liệu"),
    "btnLogin": ("fa5s.sign-in-alt", None),
    "btnExit": ("fa5s.sign-out-alt", None),
    "btnMenuDashboard": ("fa5s.home", None),
    "btnMenuSeller": ("fa5s.user-tie", None),
    "btnMenuProduct": ("fa5s.box-open", None),
    "btnMenuCustomer": ("fa5s.users", None),
    "btnMenuLivestream": ("fa5s.video", None),
    "btnMenuLivestreamDetail": ("fa5s.list-alt", None),
    "btnMenuComment": ("fa5s.comments", None),
    "btnMenuOrder": ("fa5s.shopping-cart", None),
    "btnMenuOrderDetail": ("fa5s.clipboard-list", None),
    "btnMenuPayment": ("fa5s.credit-card", None),
    "btnMenuVoucher": ("fa5s.ticket-alt", None),
    "btnMenuStatistics": ("fa5s.chart-bar", None),
    "btnMenuLogout": ("fa5s.sign-out-alt", None),
}

_EMOJI_RE = _re.compile(
    "[\U0001F000-\U0001FAFF\U00002600-\U000027BF\U0001F1E6-\U0001F1FF⬀-⯿←-⇿️⃣]+"
)


def strip_emoji(text):
    """Bỏ emoji khỏi chuỗi hiển thị (giữ nguyên chữ)."""
    return _EMOJI_RE.sub("", text or "").strip()


def apply_fontawesome_icons(root_widget):
    """Quét mọi widget con: bỏ emoji trong text, gắn icon Font Awesome cho nút đã đăng ký."""
    for w in root_widget.findChildren(QtWidgets.QWidget):
        name = w.objectName()
        if isinstance(w, QtWidgets.QAbstractButton):
            if HAS_QTA and name in _BUTTON_ICONS:
                icon_name, label = _BUTTON_ICONS[name]
                # Nút menu sidebar nằm trên nền tím đậm -> icon trắng; còn lại theo màu chữ nút
                icon_color = "#ffffff" if name.startswith("btnMenu") else w.palette().buttonText().color().name()
                try:
                    w.setIcon(qta.icon(icon_name, color=icon_color))
                except Exception:
                    w.setIcon(qta.icon(icon_name))
                w.setText(label if label is not None else strip_emoji(w.text()))
            else:
                w.setText(strip_emoji(w.text()))
        elif isinstance(w, QtWidgets.QLabel):
            if _EMOJI_RE.search(w.text() or ""):
                w.setText(strip_emoji(w.text()))
        elif isinstance(w, QtWidgets.QLineEdit):
            if _EMOJI_RE.search(w.placeholderText() or ""):
                w.setPlaceholderText(strip_emoji(w.placeholderText()))
    # Bỏ emoji trong tiêu đề cột của các bảng
    for tbl in root_widget.findChildren(QtWidgets.QTableWidget):
        for c in range(tbl.columnCount()):
            item = tbl.horizontalHeaderItem(c)
            if item and _EMOJI_RE.search(item.text() or ""):
                item.setText(strip_emoji(item.text()))


def ensure_schema():
    """Bổ sung cột LoaiUuDai/GiamToiDa cho bảng VOUCHER nếu chưa có (chạy nhiều lần vô hại)."""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(VOUCHER)")
        cols = [row[1] for row in cursor.fetchall()]
        if "LoaiUuDai" not in cols:
            cursor.execute("ALTER TABLE VOUCHER ADD COLUMN LoaiUuDai TEXT")
        if "GiamToiDa" not in cols:
            cursor.execute("ALTER TABLE VOUCHER ADD COLUMN GiamToiDa REAL")
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Lỗi ensure_schema: {e}")


def apply_order_status_effects(cursor, ma_dh, old_status, new_status):
    """Đồng bộ tồn kho + hóa đơn theo trạng thái đơn hàng.

    - Sang "Đã giao": trừ tồn kho theo chi tiết đơn, hóa đơn -> "Đã thanh toán".
    - Sang "Đã hủy": nếu trước đó đã trừ kho ("Đã giao") thì cộng trả, hóa đơn -> "Hoàn tiền".
    """
    if old_status == new_status:
        return

    if new_status == "Đã giao" and old_status != "Đã giao":
        cursor.execute("""
            UPDATE SAN_PHAM SET SoLuongTon = SoLuongTon - (
                SELECT ct.SoLuong FROM CHI_TIET_DON_HANG ct
                WHERE ct.MaDonHang = ? AND ct.MaSP = SAN_PHAM.MaSP
            )
            WHERE MaSP IN (SELECT MaSP FROM CHI_TIET_DON_HANG WHERE MaDonHang = ?)
        """, (ma_dh, ma_dh))
        cursor.execute("UPDATE HOA_DON SET TrangThaiHD = 'Đã thanh toán' WHERE MaDonHang = ?", (ma_dh,))
    elif new_status == "Đã hủy":
        if old_status == "Đã giao":
            cursor.execute("""
                UPDATE SAN_PHAM SET SoLuongTon = SoLuongTon + (
                    SELECT ct.SoLuong FROM CHI_TIET_DON_HANG ct
                    WHERE ct.MaDonHang = ? AND ct.MaSP = SAN_PHAM.MaSP
                )
                WHERE MaSP IN (SELECT MaSP FROM CHI_TIET_DON_HANG WHERE MaDonHang = ?)
            """, (ma_dh, ma_dh))
        cursor.execute("UPDATE HOA_DON SET TrangThaiHD = 'Hoàn tiền' WHERE MaDonHang = ?", (ma_dh,))


def recalc_order_total(cursor, ma_dh):
    """Tổng tiền đơn = SUM(ThanhTien) - GiaTriGiam voucher (không âm), đồng bộ sang HOA_DON."""
    # ponytail: giảm phẳng theo GiaTriGiam, bỏ qua DieuKienApDung dạng text — nâng cấp khi cần parse điều kiện
    cursor.execute("""
        UPDATE DON_HANG SET TongTien = MAX(0,
            COALESCE((SELECT SUM(ThanhTien) FROM CHI_TIET_DON_HANG WHERE MaDonHang = ?), 0)
            - COALESCE((SELECT v.GiaTriGiam FROM VOUCHER v WHERE v.MaVoucher = DON_HANG.MaVoucher), 0)
        ) WHERE MaDonHang = ?
    """, (ma_dh, ma_dh))
    cursor.execute("UPDATE HOA_DON SET TongTien = COALESCE((SELECT TongTien FROM DON_HANG WHERE MaDonHang = ?), 0) WHERE MaDonHang = ?", (ma_dh, ma_dh))


def get_widget_value(widget, widget_type):
    if widget_type == "text":
        return widget.text().strip()
    elif widget_type == "spin":
        return widget.value()
    elif widget_type == "combo_text":
        val = widget.currentText()
        if " - " in val:
            return val.split(" - ")[0]
        if val == "Không":
            return None
        return val
    elif widget_type == "combo_index":
        return widget.currentIndex()
    elif widget_type == "date":
        return widget.date().toString("yyyy-MM-dd")
    elif widget_type == "datetime":
        return widget.dateTime().toString("yyyy-MM-dd HH:mm:ss")
    return ""

def set_widget_value(widget, widget_type, value):
    val_str = str(value if value is not None else "")
    if widget_type == "text":
        widget.setText(val_str)
    elif widget_type == "spin":
        try:
            widget.setValue(int(float(value)) if value else 0)
        except Exception:
            widget.setValue(0)
    elif widget_type == "combo_text":
        if not val_str or val_str == "None":
            widget.setCurrentText("Không")
            return
        for i in range(widget.count()):
            item_text = widget.itemText(i)
            if item_text == val_str or item_text.startswith(val_str + " - "):
                widget.setCurrentIndex(i)
                return
        widget.setCurrentText(val_str)
    elif widget_type == "combo_index":
        try:
            widget.setCurrentIndex(int(value) if value else 0)
        except Exception:
            widget.setCurrentIndex(0)
    elif widget_type == "date":
        qdate = QtCore.QDate.fromString(val_str[:10], "yyyy-MM-dd")
        if qdate.isValid():
            widget.setDate(qdate)
    elif widget_type == "datetime":
        qdatetime = QtCore.QDateTime.fromString(val_str, "yyyy-MM-dd HH:mm:ss")
        if not qdatetime.isValid():
            qdatetime = QtCore.QDateTime.fromString(val_str, "yyyy-MM-dd HH:mm")
        if qdatetime.isValid():
            widget.setDateTime(qdatetime)


class DbFormController(QtCore.QObject):
    def __init__(self, main_controller, ui_page, table_widget, db_table, pk_col, col_mappings, refresh_callback):
        super().__init__()
        self.main_controller = main_controller
        self.ui = ui_page
        self.table_widget = table_widget
        self.db_table = db_table
        self.pk_col = pk_col
        self.col_mappings = col_mappings  # list of tuples: (db_col, widget, widget_type)
        self.refresh_callback = refresh_callback
        self.current_action = None
        self.selected_pk = None

        self.setup_connections()
        self.set_form_enabled(False)

    def setup_connections(self):
        # Bắt buộc chọn theo dòng để click vào ô bất kỳ cũng đổ được dữ liệu lên form
        self.table_widget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        self.table_widget.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_widget.itemSelectionChanged.connect(self.sync_table_to_form)
        self.table_widget.cellClicked.connect(lambda r, c: self.sync_table_to_form())
        if hasattr(self.ui, 'btnThem'):
            self.ui.btnThem.clicked.connect(self.action_them)
        if hasattr(self.ui, 'btnSua'):
            self.ui.btnSua.clicked.connect(self._on_edit_button)
        if hasattr(self.ui, 'btnXoa'):
            self.ui.btnXoa.clicked.connect(self.action_xoa)
        if hasattr(self.ui, 'btnLuu'):
            self.ui.btnLuu.clicked.connect(self.action_luu)
        if hasattr(self.ui, 'btnHuy'):
            self.ui.btnHuy.clicked.connect(self.action_huy)
        if hasattr(self.ui, 'btnRefresh'):
            self.ui.btnRefresh.clicked.connect(self.action_refresh)
        if hasattr(self.ui, 'btnSearch'):
            self.ui.btnSearch.clicked.connect(self.action_search)
        if hasattr(self.ui, 'btnXuatExcel'):
            self.ui.btnXuatExcel.clicked.connect(self.action_xuat_excel)

    def set_form_enabled(self, enabled=True):
        for db_col, widget, widget_type in self.col_mappings:
            if widget:
                widget.setEnabled(enabled)
        
        if hasattr(self.ui, 'btnThem'):
            self.ui.btnThem.setEnabled(not enabled)
        if hasattr(self.ui, 'btnXoa'):
            self.ui.btnXoa.setEnabled(not enabled)
        # Nút Sửa toggle thành Lưu: đang sửa/thêm -> "Lưu", ngược lại -> "Sửa"
        if hasattr(self.ui, 'btnSua'):
            self.ui.btnSua.setEnabled(True)
            if enabled:
                self.ui.btnSua.setText("Lưu")
                if HAS_QTA:
                    self.ui.btnSua.setIcon(qta.icon('fa5s.save', color='white'))
            else:
                self.ui.btnSua.setText("Sửa")
                if HAS_QTA:
                    self.ui.btnSua.setIcon(qta.icon('fa5s.edit', color='white'))
        # Ẩn nút Lưu riêng — mọi submit đi qua nút Sửa/Lưu đã toggle
        if hasattr(self.ui, 'btnLuu'):
            self.ui.btnLuu.setVisible(False)
        if hasattr(self.ui, 'btnHuy'):
            self.ui.btnHuy.setEnabled(enabled)

    def sync_table_to_form(self):
        if self.current_action is not None:
            return
        
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            self.selected_pk = None
            return
            
        row = selected_items[0].row()
        self.selected_pk = self.table_widget.item(row, 0).text()
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            if self.db_table == "LIVESTREAM_SAN_PHAM":
                parts = self.selected_pk.split("-")
                if len(parts) >= 2:
                    cursor.execute("""
                        SELECT lsp.MaLive, lsp.MaSP, p.SoLuongTon, p.GiaBan
                        FROM LIVESTREAM_SAN_PHAM lsp
                        JOIN SAN_PHAM p ON lsp.MaSP = p.MaSP
                        WHERE lsp.MaLive = ? AND lsp.MaSP = ?
                    """, (parts[0], parts[1]))
                    row_data = cursor.fetchone()
                else:
                    row_data = None
            elif self.db_table == "CHI_TIET_DON_HANG":
                parts = self.selected_pk.split("-")
                if len(parts) >= 2:
                    cursor.execute("SELECT MaDonHang, MaSP, SoLuong, DonGia, ThanhTien FROM CHI_TIET_DON_HANG WHERE MaDonHang = ? AND MaSP = ?", (parts[0], parts[1]))
                    row_data = cursor.fetchone()
                else:
                    row_data = None
            elif self.db_table == "DON_HANG":
                cursor.execute("""
                    SELECT 
                        dh.NgayDat,
                        bl.MaLive,
                        kh.HoTen,
                        kh.SoDienThoai,
                        kh.DiaChi,
                        dh.MaVoucher,
                        dh.TrangThaiDH,
                        dh.TongTien
                    FROM DON_HANG dh
                    LEFT JOIN KHACH_HANG kh ON dh.MaKhachHang = kh.MaKhachHang
                    LEFT JOIN BINH_LUAN bl ON dh.MaBinhLuan = bl.MaBinhLuan
                    WHERE dh.MaDonHang = ?
                """, (self.selected_pk,))
                order_res = cursor.fetchone()
                if order_res:
                    row_data = (order_res[0], order_res[1], order_res[2], order_res[3], order_res[4], order_res[5], order_res[6], order_res[7])
                else:
                    row_data = None
                
                # Fetch order products (kèm ảnh sản phẩm nếu có trong thư mục dự án)
                cursor.execute("""
                    SELECT p.TenSP, ct.SoLuong, ct.DonGia, p.HinhAnh
                    FROM CHI_TIET_DON_HANG ct
                    JOIN SAN_PHAM p ON ct.MaSP = p.MaSP
                    WHERE ct.MaDonHang = ?
                """, (self.selected_pk,))
                products = cursor.fetchall()
                if hasattr(self.ui, "tblOrderProductsList"):
                    self.main_controller.populate_table(self.ui.tblOrderProductsList,
                                                        [p[:3] for p in products])
                    # Gắn icon ảnh vào cột tên sản phẩm
                    for r, prod in enumerate(products):
                        img = prod[3]
                        item = self.ui.tblOrderProductsList.item(r, 0)
                        if not (img and item):
                            continue
                        for candidate in (os.path.join(current_dir, img),
                                          os.path.join(current_dir, os.path.splitext(img)[0] + ".png")):
                            if os.path.exists(candidate):
                                item.setIcon(QtGui.QIcon(candidate))
                                self.ui.tblOrderProductsList.setIconSize(QtCore.QSize(32, 32))
                                break
            else:
                cols = [col for col, _, _ in self.col_mappings if col is not None]
                query = f"SELECT {', '.join(cols)} FROM {self.db_table} WHERE {self.pk_col} = ?"
                cursor.execute(query, (self.selected_pk,))
                row_data = cursor.fetchone()
            conn.close()
            
            if row_data:
                for idx, (db_col, widget, widget_type) in enumerate(self.col_mappings):
                    if widget:
                        set_widget_value(widget, widget_type, row_data[idx])
        except Exception as e:
            print(f"Error syncing {self.db_table} to form: {e}")

    def get_pk_widget(self):
        for db_col, widget, widget_type in self.col_mappings:
            if db_col == self.pk_col:
                return widget
        return None

    def get_pk_widget_type(self):
        for db_col, widget, widget_type in self.col_mappings:
            if db_col == self.pk_col:
                return widget_type
        return "text"

    def clear_form(self):
        for db_col, widget, widget_type in self.col_mappings:
            if widget:
                if widget_type == "text":
                    widget.clear()
                elif widget_type == "combo_text":
                    widget.setCurrentIndex(0)
                elif widget_type == "combo_index":
                    widget.setCurrentIndex(0)
                elif widget_type == "spin":
                    widget.setValue(0)
                elif widget_type == "date":
                    widget.setDate(QtCore.QDate.currentDate())
                elif widget_type == "datetime":
                    widget.setDateTime(QtCore.QDateTime.currentDateTime())

    def generate_new_pk(self):
        prefix = ""
        if self.db_table == "KHACH_HANG": prefix = "KH"
        elif self.db_table == "SAN_PHAM": prefix = "SP"
        elif self.db_table == "NGUOI_BAN": prefix = "NB"
        elif self.db_table == "LIVESTREAM": prefix = "LIVE"
        elif self.db_table == "DON_HANG": prefix = "DH"
        elif self.db_table == "BINH_LUAN": prefix = "BL"
        elif self.db_table == "VOUCHER": prefix = "VOUCHER"
        elif self.db_table == "HOA_DON": prefix = "HD"
        return self.generate_new_pk_generic(self.db_table, self.pk_col, prefix)

    def generate_new_pk_generic(self, db_table, pk_col, prefix):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(f"SELECT {pk_col} FROM {db_table}")
            pks = [row[0] for row in cursor.fetchall() if row[0] and row[0].startswith(prefix)]
            conn.close()
            
            max_num = 0
            for pk in pks:
                num_part = pk[len(prefix):]
                if num_part.isdigit():
                    max_num = max(max_num, int(num_part))
            
            return f"{prefix}{max_num + 1:02d}"
        except Exception as e:
            print(f"Error generating PK: {e}")
            return prefix + "99"

    def _on_edit_button(self):
        # Nút Sửa/Lưu gộp: đang mở form (THEM/SUA) -> submit; ngược lại -> mở sửa
        if self.current_action in ("THEM", "SUA"):
            self.action_luu()
        else:
            self.action_sua()

    def action_them(self):
        self.current_action = "THEM"
        self.selected_pk = None
        self.clear_form()
        self.set_form_enabled(True)
        pk_widget = self.get_pk_widget()
        if pk_widget:
            pk_widget.setEnabled(True)

    def action_sua(self):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.main_controller, "Cảnh báo", "Vui lòng chọn một bản ghi trên bảng để sửa!")
            return
        self.current_action = "SUA"
        self.set_form_enabled(True)
        pk_widget = self.get_pk_widget()
        if pk_widget:
            pk_widget.setEnabled(False)

    def action_xoa(self):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.main_controller, "Cảnh báo", "Vui lòng chọn một bản ghi trên bảng để xóa!")
            return
            
        row = selected_items[0].row()
        pk_val = self.table_widget.item(row, 0).text()
        
        confirm = QMessageBox.question(
            self.main_controller, "Xác nhận xóa",
            f"Bạn có chắc chắn muốn xóa bản ghi {pk_val} không?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if confirm == QMessageBox.StandardButton.Yes:
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                
                if self.db_table == "LIVESTREAM_SAN_PHAM":
                    parts = pk_val.split("-")
                    cursor.execute("DELETE FROM LIVESTREAM_SAN_PHAM WHERE MaLive = ? AND MaSP = ?", (parts[0], parts[1]))
                elif self.db_table == "CHI_TIET_DON_HANG":
                    parts = pk_val.split("-")
                    cursor.execute("DELETE FROM CHI_TIET_DON_HANG WHERE MaDonHang = ? AND MaSP = ?", (parts[0], parts[1]))

                    # Update Totals (đã trừ voucher nếu có)
                    recalc_order_total(cursor, parts[0])
                elif self.db_table == "DON_HANG":
                    # Delete invoices and order details first to avoid constraint violation
                    cursor.execute("DELETE FROM HOA_DON WHERE MaDonHang = ?", (pk_val,))
                    cursor.execute("DELETE FROM CHI_TIET_DON_HANG WHERE MaDonHang = ?", (pk_val,))
                    cursor.execute(f"DELETE FROM {self.db_table} WHERE {self.pk_col} = ?", (pk_val,))
                else:
                    cursor.execute(f"DELETE FROM {self.db_table} WHERE {self.pk_col} = ?", (pk_val,))
                
                conn.commit()
                conn.close()
                self.refresh_callback()
                self.clear_form()
                # Đồng bộ các màn hình liên quan (bảng, combobox, thẻ số liệu, biểu đồ)
                if hasattr(self.main_controller, 'refresh_dependents'):
                    self.main_controller.refresh_dependents()
                QMessageBox.information(self.main_controller, "Thành công", f"Đã xóa thành công bản ghi {pk_val}.")
            except Exception as e:
                QMessageBox.critical(self.main_controller, "Lỗi", f"Không thể xóa bản ghi: {e}")

    def action_luu(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            if self.db_table == "LIVESTREAM_SAN_PHAM":
                ma_live = get_widget_value(self.ui.cbLivestream, "combo_text")
                ma_sp = get_widget_value(self.ui.cbSanPham, "combo_text")
                if not ma_live or not ma_sp:
                    QMessageBox.warning(self.main_controller, "Lỗi dữ liệu", "Vui lòng chọn livestream và sản phẩm!")
                    conn.close()
                    return
                
                if self.current_action == "THEM":
                    cursor.execute("SELECT 1 FROM LIVESTREAM_SAN_PHAM WHERE MaLive = ? AND MaSP = ?", (ma_live, ma_sp))
                    if cursor.fetchone():
                        QMessageBox.warning(self.main_controller, "Lỗi", "Bản ghi liên kết này đã tồn tại!")
                        conn.close()
                        return
                    cursor.execute("INSERT INTO LIVESTREAM_SAN_PHAM (MaLive, MaSP) VALUES (?, ?)", (ma_live, ma_sp))
                elif self.current_action == "SUA":
                    parts = self.selected_pk.split("-")
                    cursor.execute("DELETE FROM LIVESTREAM_SAN_PHAM WHERE MaLive = ? AND MaSP = ?", (parts[0], parts[1]))
                    cursor.execute("INSERT INTO LIVESTREAM_SAN_PHAM (MaLive, MaSP) VALUES (?, ?)", (ma_live, ma_sp))
                
                conn.commit()
                conn.close()
                self.current_action = None
                self.set_form_enabled(False)
                self.refresh_callback()
                if hasattr(self.main_controller, 'refresh_dependents'):
                    self.main_controller.refresh_dependents()
                QMessageBox.information(self.main_controller, "Thành công", "Lưu liên kết livestream - sản phẩm thành công.")
                return

            elif self.db_table == "CHI_TIET_DON_HANG":
                ma_dh = get_widget_value(self.ui.cbDonHang, "combo_text")
                ma_sp = get_widget_value(self.ui.cbSanPham, "combo_text")
                so_luong = get_widget_value(self.ui.spinSoLuong, "spin")
                don_gia_str = self.ui.txtGiaBan.text().strip().replace(",", "")
                
                if not ma_dh or not ma_sp:
                    QMessageBox.warning(self.main_controller, "Lỗi dữ liệu", "Vui lòng chọn đơn hàng và sản phẩm!")
                    conn.close()
                    return
                try:
                    don_gia = float(don_gia_str) if don_gia_str else 0.0
                except ValueError:
                    QMessageBox.warning(self.main_controller, "Lỗi dữ liệu", "Giá bán phải là số hợp lệ!")
                    conn.close()
                    return
                
                thanh_tien = so_luong * don_gia
                
                if self.current_action == "THEM":
                    cursor.execute("SELECT 1 FROM CHI_TIET_DON_HANG WHERE MaDonHang = ? AND MaSP = ?", (ma_dh, ma_sp))
                    if cursor.fetchone():
                        QMessageBox.warning(self.main_controller, "Lỗi", "Chi tiết đơn hàng này đã tồn tại!")
                        conn.close()
                        return
                    cursor.execute("INSERT INTO CHI_TIET_DON_HANG (MaDonHang, MaSP, SoLuong, DonGia, ThanhTien) VALUES (?, ?, ?, ?, ?)",
                                   (ma_dh, ma_sp, so_luong, don_gia, thanh_tien))
                elif self.current_action == "SUA":
                    parts = self.selected_pk.split("-")
                    cursor.execute("DELETE FROM CHI_TIET_DON_HANG WHERE MaDonHang = ? AND MaSP = ?", (parts[0], parts[1]))
                    cursor.execute("INSERT INTO CHI_TIET_DON_HANG (MaDonHang, MaSP, SoLuong, DonGia, ThanhTien) VALUES (?, ?, ?, ?, ?)",
                                   (ma_dh, ma_sp, so_luong, don_gia, thanh_tien))
                
                # Update Totals on DON_HANG and HOA_DON (đã trừ voucher nếu có)
                recalc_order_total(cursor, ma_dh)
                
                conn.commit()
                conn.close()
                self.current_action = None
                self.set_form_enabled(False)
                self.refresh_callback()
                if hasattr(self.main_controller, 'refresh_dependents'):
                    self.main_controller.refresh_dependents()
                QMessageBox.information(self.main_controller, "Thành công", "Lưu chi tiết đơn hàng thành công.")
                return

            elif self.db_table == "DON_HANG":
                ten_kh = self.ui.txtTenKhachHang.text().strip()
                sdt = self.ui.txtSoDienThoai.text().strip()
                dia_chi = self.ui.txtDiaChiGiao.text().strip()
                ngay_dat = self.ui.dateNgayDat.dateTime().toString("yyyy-MM-dd HH:mm:ss")
                ma_voucher = get_widget_value(self.ui.cbVoucher, "combo_text")
                trang_thai = self.ui.cbTrangThai.currentText()
                pt_tt = self.ui.cbPaymentMode.currentText()
                
                if not ten_kh or not sdt:
                    QMessageBox.warning(self.main_controller, "Lỗi dữ liệu", "Vui lòng nhập tên và số điện thoại khách hàng!")
                    conn.close()
                    return
                
                # Check / Insert Customer
                cursor.execute("SELECT MaKhachHang FROM KHACH_HANG WHERE SoDienThoai = ?", (sdt,))
                cust_res = cursor.fetchone()
                if cust_res:
                    ma_kh = cust_res[0]
                    cursor.execute("UPDATE KHACH_HANG SET HoTen = ?, DiaChi = ? WHERE MaKhachHang = ?", (ten_kh, dia_chi, ma_kh))
                else:
                    ma_kh = self.generate_new_pk_generic("KHACH_HANG", "MaKhachHang", "KH")
                    cursor.execute("INSERT INTO KHACH_HANG (MaKhachHang, HoTen, SoDienThoai, Email, DiaChi) VALUES (?, ?, ?, ?, ?)",
                                   (ma_kh, ten_kh, sdt, "", dia_chi))
                
                # Tổng tiền nhập tay (chi tiết đơn hàng sẽ ghi đè khi đơn có sản phẩm)
                tong_tien_str = self.ui.txtTongTien.text().strip().replace(",", "") if hasattr(self.ui, 'txtTongTien') else ""
                try:
                    tong_tien_manual = float(tong_tien_str) if tong_tien_str else None
                except ValueError:
                    tong_tien_manual = None

                old_status = None
                if self.current_action == "THEM":
                    ma_dh = self.generate_new_pk_generic("DON_HANG", "MaDonHang", "DH")
                    cursor.execute("INSERT INTO DON_HANG (MaDonHang, MaKhachHang, MaVoucher, NgayDat, TongTien, TrangThaiDH) VALUES (?, ?, ?, ?, ?, ?)",
                                   (ma_dh, ma_kh, ma_voucher, ngay_dat, tong_tien_manual or 0.0, trang_thai))

                    # Create corresponding Invoice
                    ma_hd = self.generate_new_pk_generic("HOA_DON", "MaHoaDon", "HD")
                    cursor.execute("INSERT INTO HOA_DON (MaHoaDon, MaDonHang, PhuongThucTT, TongTien, ThoiGianLap, TrangThaiHD) VALUES (?, ?, ?, ?, ?, 'Chưa thanh toán')",
                                   (ma_hd, ma_dh, pt_tt, tong_tien_manual or 0.0, ngay_dat))
                elif self.current_action == "SUA":
                    ma_dh = self.selected_pk
                    cursor.execute("SELECT TrangThaiDH FROM DON_HANG WHERE MaDonHang = ?", (ma_dh,))
                    res = cursor.fetchone()
                    old_status = res[0] if res else None
                    cursor.execute("UPDATE DON_HANG SET MaKhachHang = ?, MaVoucher = ?, NgayDat = ?, TrangThaiDH = ? WHERE MaDonHang = ?",
                                   (ma_kh, ma_voucher, ngay_dat, trang_thai, ma_dh))
                    if tong_tien_manual is not None:
                        cursor.execute("UPDATE DON_HANG SET TongTien = ? WHERE MaDonHang = ?", (tong_tien_manual, ma_dh))
                    cursor.execute("UPDATE HOA_DON SET PhuongThucTT = ? WHERE MaDonHang = ?", (pt_tt, ma_dh))

                # Đồng bộ tồn kho + trạng thái hóa đơn theo trạng thái đơn
                apply_order_status_effects(cursor, ma_dh, old_status, trang_thai)
                if trang_thai not in ("Đã giao", "Đã hủy"):
                    cursor.execute("UPDATE HOA_DON SET TrangThaiHD = 'Chưa thanh toán' WHERE MaDonHang = ?", (ma_dh,))

                # Đơn có chi tiết thì tổng tiền tính từ chi tiết (trừ voucher)
                cursor.execute("SELECT 1 FROM CHI_TIET_DON_HANG WHERE MaDonHang = ? LIMIT 1", (ma_dh,))
                if cursor.fetchone():
                    recalc_order_total(cursor, ma_dh)

                conn.commit()
                conn.close()
                self.current_action = None
                self.set_form_enabled(False)
                self.refresh_callback()
                if hasattr(self.main_controller, 'refresh_dependents'):
                    self.main_controller.refresh_dependents()
                QMessageBox.information(self.main_controller, "Thành công", "Lưu thông tin đơn hàng thành công.")
                return

            else:
                # Standard save flow
                vals = {}
                for db_col, widget, widget_type in self.col_mappings:
                    if widget:
                        vals[db_col] = get_widget_value(widget, widget_type)
                
                # Combo voucher bên Đơn hàng hiển thị theo TenVoucher — mặc định dùng mã
                if self.db_table == "VOUCHER" and not vals.get("TenVoucher"):
                    vals["TenVoucher"] = vals.get("MaVoucher") or self.selected_pk

                pk_widget = self.get_pk_widget()
                if self.current_action == "THEM":
                    if pk_widget:
                        pk_val = get_widget_value(pk_widget, self.get_pk_widget_type())
                        if not pk_val:
                            QMessageBox.warning(self.main_controller, "Lỗi dữ liệu", "Vui lòng nhập Mã khóa chính!")
                            conn.close()
                            return
                        vals[self.pk_col] = pk_val
                    else:
                        pk_val = self.generate_new_pk()
                        vals[self.pk_col] = pk_val
                    
                    cols = list(vals.keys())
                    placeholders = [f":" + c for c in cols]
                    query = f"INSERT INTO {self.db_table} ({', '.join(cols)}) VALUES ({', '.join(placeholders)})"
                    cursor.execute(query, vals)
                    
                elif self.current_action == "SUA":
                    if pk_widget:
                        pk_val = get_widget_value(pk_widget, self.get_pk_widget_type())
                    else:
                        pk_val = self.selected_pk
                        
                    if not pk_val:
                        QMessageBox.warning(self.main_controller, "Lỗi", "Không xác định được Mã bản ghi cần cập nhật!")
                        conn.close()
                        return
                        
                    set_clause = ", ".join([f"{col} = :{col}" for col in vals.keys() if col != self.pk_col])
                    vals[self.pk_col] = pk_val
                    query = f"UPDATE {self.db_table} SET {set_clause} WHERE {self.pk_col} = :{self.pk_col}"
                    cursor.execute(query, vals)
                
                conn.commit()
                conn.close()
                self.current_action = None
                self.set_form_enabled(False)
                self.refresh_callback()
                if hasattr(self.main_controller, 'refresh_dependents'):
                    self.main_controller.refresh_dependents()
                QMessageBox.information(self.main_controller, "Thành công", "Lưu dữ liệu thành công.")
        except Exception as e:
            QMessageBox.critical(self.main_controller, "Lỗi", f"Không thể lưu dữ liệu: {e}")

    def action_huy(self):
        self.current_action = None
        self.clear_form()
        self.set_form_enabled(False)
        self.selected_pk = None
        # Nạp lại bảng thay vì xóa trắng để không gây cảm giác mất dữ liệu
        self.refresh_callback()
        if hasattr(self.ui, "lblStatus"):
            self.ui.lblStatus.setText("Trạng thái: Đã hủy thao tác.")

    def action_refresh(self):
        if hasattr(self.ui, 'txtSearch'):
            self.ui.txtSearch.clear()
        self.refresh_callback()

    def action_search(self):
        if not hasattr(self.ui, 'txtSearch'):
            return
        search_text = self.ui.txtSearch.text().strip()
        if not search_text:
            self.refresh_callback()
            return
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            if self.db_table == "BINH_LUAN":
                query = """
                    SELECT MaBinhLuan, MaLive, NguoiBinhLuan, NoiDung, ThoiGian 
                    FROM BINH_LUAN 
                    WHERE NguoiBinhLuan LIKE ? OR NoiDung LIKE ? OR MaLive LIKE ?
                """
                cursor.execute(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
            elif self.db_table == "KHACH_HANG":
                query = """
                    SELECT MaKhachHang, HoTen, SoDienThoai, Email, DiaChi, '2026-07-15' 
                    FROM KHACH_HANG 
                    WHERE HoTen LIKE ? OR SoDienThoai LIKE ? OR Email LIKE ? OR DiaChi LIKE ?
                """
                cursor.execute(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
            elif self.db_table == "LIVESTREAM":
                query = """
                    SELECT MaLive, TenLive, ThoiGianBatDau, MaNguoiBan, TrangThai 
                    FROM LIVESTREAM 
                    WHERE TenLive LIKE ? OR MaLive LIKE ? OR MaNguoiBan LIKE ? OR TrangThai LIKE ?
                """
                cursor.execute(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
            elif self.db_table == "LIVESTREAM_SAN_PHAM":
                query = """
                    SELECT 
                        (lsp.MaLive || '-' || lsp.MaSP) AS MaChiTietLive,
                        l.TenLive,
                        lsp.MaSP,
                        p.TenSP,
                        p.SoLuongTon,
                        p.GiaBan
                    FROM LIVESTREAM_SAN_PHAM lsp
                    JOIN LIVESTREAM l ON lsp.MaLive = l.MaLive
                    JOIN SAN_PHAM p ON lsp.MaSP = p.MaSP
                    WHERE l.TenLive LIKE ? OR p.TenSP LIKE ? OR lsp.MaLive LIKE ? OR lsp.MaSP LIKE ?
                """
                cursor.execute(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
            elif self.db_table == "DON_HANG":
                query = """
                    SELECT 
                        dh.MaDonHang,
                        dh.NgayDat,
                        kh.HoTen,
                        kh.SoDienThoai,
                        dh.TongTien,
                        COALESCE(v.TenVoucher, dh.MaVoucher),
                        dh.TrangThaiDH,
                        COALESCE((SELECT GROUP_CONCAT(p.TenSP, ', ')
                                  FROM CHI_TIET_DON_HANG ct JOIN SAN_PHAM p ON ct.MaSP = p.MaSP
                                  WHERE ct.MaDonHang = dh.MaDonHang), '(chưa có SP)')
                    FROM DON_HANG dh
                    LEFT JOIN KHACH_HANG kh ON dh.MaKhachHang = kh.MaKhachHang
                    LEFT JOIN VOUCHER v ON dh.MaVoucher = v.MaVoucher
                    WHERE dh.MaDonHang LIKE ? OR kh.HoTen LIKE ? OR kh.SoDienThoai LIKE ? OR dh.TrangThaiDH LIKE ?
                """
                cursor.execute(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
            elif self.db_table == "CHI_TIET_DON_HANG":
                query = """
                    SELECT 
                        (ct.MaDonHang || '-' || ct.MaSP) AS MaChiTietDH,
                        ct.MaDonHang,
                        ct.MaSP,
                        p.TenSP,
                        ct.SoLuong,
                        ct.DonGia,
                        ct.ThanhTien,
                        COALESCE(v.TenVoucher, dh.MaVoucher, 'Không')
                    FROM CHI_TIET_DON_HANG ct
                    JOIN SAN_PHAM p ON ct.MaSP = p.MaSP
                    LEFT JOIN DON_HANG dh ON ct.MaDonHang = dh.MaDonHang
                    LEFT JOIN VOUCHER v ON dh.MaVoucher = v.MaVoucher
                    WHERE ct.MaDonHang LIKE ? OR ct.MaSP LIKE ? OR p.TenSP LIKE ?
                """
                cursor.execute(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
            elif self.db_table == "HOA_DON":
                query = """
                    SELECT MaHoaDon, MaDonHang, ThoiGianLap, TongTien, PhuongThucTT, TrangThaiHD 
                    FROM HOA_DON 
                    WHERE MaHoaDon LIKE ? OR MaDonHang LIKE ? OR PhuongThucTT LIKE ? OR TrangThaiHD LIKE ?
                """
                cursor.execute(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
            elif self.db_table == "VOUCHER":
                query = """
                    SELECT
                        MaVoucher,
                        COALESCE(LoaiUuDai, 'Số tiền cố định (VND)') AS LoaiUuDai,
                        GiaTriGiam,
                        DieuKienApDung,
                        COALESCE(GiamToiDa, GiaTriGiam) AS GiamToiDa,
                        NgayBatDau,
                        NgayKetThuc,
                        TrangThai
                    FROM VOUCHER
                    WHERE MaVoucher LIKE ? OR TenVoucher LIKE ? OR TrangThai LIKE ?
                """
                cursor.execute(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
            else:
                conn.close()
                self.main_controller.search_table_generic(self.table_widget, self.db_table, self.pk_col, search_text, self.refresh_callback)
                return
            
            results = cursor.fetchall()
            conn.close()
            self.main_controller.populate_table(self.table_widget, results)
            if hasattr(self.ui, 'lblStatus'):
                self.ui.lblStatus.setText(f"Trạng thái: Tìm thấy {len(results)} kết quả.")
        except Exception as e:
            print(f"Error action_search for {self.db_table}: {e}")

    def action_xuat_excel(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self.main_controller, "Lưu file Excel", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = self.db_table[:30]

            headers = []
            for col in range(self.table_widget.columnCount()):
                header_item = self.table_widget.horizontalHeaderItem(col)
                headers.append(header_item.text() if header_item else f"Cột {col+1}")
            ws.append(headers)

            for row in range(self.table_widget.rowCount()):
                row_data = []
                for col in range(self.table_widget.columnCount()):
                    item = self.table_widget.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            wb.save(path)
            QMessageBox.information(self.main_controller, "Thành công", f"Đã xuất dữ liệu Excel thành công tại:\n{path}")
        except ImportError:
            QMessageBox.critical(self.main_controller, "Lỗi", "Vui lòng cài đặt openpyxl bằng lệnh: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self.main_controller, "Lỗi", f"Không thể lưu file Excel: {e}")


class DashboardEx(QMainWindow):
    def __init__(self, seller_id="Admin", seller_name="Người bán", login_window=None):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.seller_id = seller_id
        self.seller_name = seller_name
        self.login_window = login_window

        # Hiển thị tên người bán đăng nhập thành công lên Header
        if hasattr(self.ui, 'lblUserScope'):
            self.ui.lblUserScope.setText(f"Xin chào, {self.seller_name} ({self.seller_id})")

        # 0. Bổ sung cột còn thiếu cho schema (an toàn khi chạy lại)
        ensure_schema()

        # 1. Khởi tạo các trang con
        self.init_sub_pages()

        # 2. Định tuyến Menu Sidebar và Nút Đăng xuất
        self.connect_sidebar_buttons()

        # 2b. Tinh chỉnh form: mở khóa ô nhập, thêm cột bảng, khu vực biểu đồ
        self.setup_form_tweaks()
        self.setup_charts_area()
        self.refactor_statistics_ui()

        # 3. Đọc dữ liệu từ SQLite đẩy lên Dashboard
        self.load_all_data_from_database()

        # 4. Thiết lập sự kiện và chức năng cho màn hình Người bán & Sản phẩm
        self.setup_seller_product_actions()

        # 5. Thiết lập sự kiện và chức năng cho tất cả các giao diện khác
        self.setup_all_other_actions()

        # 6. Vẽ biểu đồ lần đầu + thay emoji bằng icon Font Awesome
        self.update_all_charts()
        apply_fontawesome_icons(self)
        self.setup_tab_badges()

    def setup_tab_badges(self):
        """Gắn icon Font Awesome (trắng) vào badge tròn gradient ở đầu mỗi tab."""
        if not HAS_QTA:
            return
        badges = {
            "ui_seller": "fa5s.user-tie",
            "ui_product": "fa5s.box-open",
            "ui_customer": "fa5s.users",
            "ui_livestream": "fa5s.video",
            "ui_livestream_detail": "fa5s.list-alt",
            "ui_comment": "fa5s.comments",
            "ui_order": "fa5s.shopping-cart",
            "ui_order_detail": "fa5s.clipboard-list",
            "ui_payment": "fa5s.credit-card",
            "ui_voucher": "fa5s.ticket-alt",
            "ui_statistics": "fa5s.chart-bar",
        }
        for ui_attr, icon_name in badges.items():
            ui_page = getattr(self, ui_attr, None)
            badge = getattr(ui_page, "lblIconBadge", None) if ui_page else None
            if badge is None:
                continue
            badge.setText("")
            pix = qta.icon(icon_name, color="#ffffff").pixmap(QtCore.QSize(24, 24))
            badge.setPixmap(pix)
            badge.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

        # Logo thương hiệu ở sidebar (nền gradient tròn) — biểu tượng phát trực tiếp
        if hasattr(self.ui, "lblSidebarLogo"):
            self.ui.lblSidebarLogo.setText("")
            self.ui.lblSidebarLogo.setPixmap(qta.icon("fa5s.broadcast-tower", color="#ffffff").pixmap(QtCore.QSize(30, 30)))
            self.ui.lblSidebarLogo.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

        # Badge icon tròn màu cho các thẻ số liệu tổng quan trên Dashboard
        cards = {
            "lblCardCustomerIcon": ("fa5s.users", "#8b5cf6"),
            "lblCardLivestreamIcon": ("fa5s.video", "#ec4899"),
            "lblCardProductIcon": ("fa5s.box-open", "#f59e0b"),
            "lblCardOrderIcon": ("fa5s.shopping-cart", "#3b82f6"),
            "lblCardRevenueIcon": ("fa5s.coins", "#10b981"),
            "lblCardVoucherIcon": ("fa5s.ticket-alt", "#ef4444"),
        }
        for lbl_name, (icon_name, color) in cards.items():
            lbl = getattr(self.ui, lbl_name, None)
            if lbl is None:
                continue
            lbl.setText("")
            lbl.setMinimumSize(QtCore.QSize(44, 44))
            lbl.setMaximumSize(QtCore.QSize(44, 44))
            lbl.setPixmap(qta.icon(icon_name, color="#ffffff").pixmap(QtCore.QSize(22, 22)))
            lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet(f"background-color: {color}; border-radius: 22px;")

        self.style_dashboard_panels()

    def style_dashboard_panels(self):
        """Bọc nền trắng bo góc cho thẻ số liệu và 2 bảng, đồng bộ với 2 khung biểu đồ."""
        panel_css = ("background-color: #ffffff; border: 1px solid #e2e8f0;"
                     " border-radius: 12px;")
        # 6 thẻ số liệu tổng quan
        for card_name in ("cardCustomer", "cardLivestream", "cardProduct",
                          "cardOrder", "cardRevenue", "cardVoucher"):
            card = getattr(self.ui, card_name, None)
            if card is not None:
                # icon con đã có style riêng nên chỉ nhắm đúng QFrame gốc
                card.setStyleSheet(f"QFrame#{card_name} {{ {panel_css} }}")

        # 2 bảng bên dưới: bảng trong suốt để lộ nền card, header không kẻ viền
        for tbl in (self.ui.tblTopProducts, self.ui.tblActiveLivestreams):
            tbl.setStyleSheet(
                "QTableWidget { background-color: transparent; border: none;"
                " gridline-color: #eef2f7; }"
                "QHeaderView::section { background-color: #f8fafc; color: #334155;"
                " border: none; padding: 6px; font-weight: 600; }"
            )

        # Gộp tiêu đề + bảng vào chung một card nền trắng (giống khung biểu đồ)
        self._wrap_in_card(self.ui.verticalLayout_table1)
        self._wrap_in_card(self.ui.verticalLayout_table2)

        # In đậm tất cả tiêu đề panel trên Dashboard cho dễ nhận biết
        for lbl_name in ("lblChart1Title", "lblChart2Title", "lblTable1Header", "lblTable2Header"):
            lbl = getattr(self.ui, lbl_name, None)
            if lbl is not None:
                f = lbl.font()
                f.setBold(True)
                f.setPointSize(max(f.pointSize(), 12))
                lbl.setFont(f)
                lbl.setStyleSheet("color: #1e293b;")

    def _wrap_in_card(self, inner_layout):
        """Chuyển một QVBoxLayout (tiêu đề + bảng) vào trong QFrame nền trắng bo góc."""
        idx = self.ui.layoutTables.indexOf(inner_layout)
        if idx < 0:
            return
        # Gỡ layout con khỏi layoutTables để có thể gắn vào frame (layout không được có 2 parent)
        self.ui.layoutTables.removeItem(inner_layout)
        card = QtWidgets.QFrame(parent=self.ui.pageDashboard)
        card.setObjectName(f"panelCard{idx}")
        card.setStyleSheet(f"QFrame#panelCard{idx} {{ background-color: #ffffff;"
                           " border: 1px solid #e2e8f0; border-radius: 12px; }")
        card.setLayout(inner_layout)  # reparent layout + các widget con vào frame
        inner_layout.setContentsMargins(16, 14, 16, 16)
        self.ui.layoutTables.insertWidget(idx, card)

    def setup_form_tweaks(self):
        """Mở khóa các ô nhập bị readOnly trong .ui và bổ sung cột bảng bằng code."""
        # Ô Tổng tiền (Order) và Thành tiền (OrderDetail) bị đặt readOnly trong Qt Designer
        self.ui_order.txtTongTien.setReadOnly(False)
        self.ui_order_detail.txtThanhTien.setReadOnly(False)

        # Bảng Đơn hàng: thêm cột "Sản phẩm" hiển thị các SP khách mua
        tbl = self.ui_order.tblOrder
        col = tbl.columnCount()
        tbl.setColumnCount(col + 1)
        tbl.setHorizontalHeaderItem(col, QTableWidgetItem("Sản phẩm"))

        # Bảng Chi tiết đơn: thêm cột "Voucher" của đơn hàng tương ứng
        tbl = self.ui_order_detail.tblOrderDetail
        col = tbl.columnCount()
        tbl.setColumnCount(col + 1)
        tbl.setHorizontalHeaderItem(col, QTableWidgetItem("Voucher"))

        # OrderDetail: chọn sản phẩm tự điền giá bán; SL/giá đổi thì tự tính thành tiền
        self.ui_order_detail.cbSanPham.currentIndexChanged.connect(self._orderdetail_autofill_price)
        self.ui_order_detail.spinSoLuong.valueChanged.connect(self._orderdetail_recalc_total)
        self.ui_order_detail.txtGiaBan.textEdited.connect(self._orderdetail_recalc_total)

    def _orderdetail_autofill_price(self):
        ma_sp = get_widget_value(self.ui_order_detail.cbSanPham, "combo_text")
        if not ma_sp:
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT GiaBan FROM SAN_PHAM WHERE MaSP = ?", (ma_sp,))
            row = cursor.fetchone()
            conn.close()
            if row and row[0] is not None:
                self.ui_order_detail.txtGiaBan.setText(f"{row[0]:g}")
                self._orderdetail_recalc_total()
        except Exception:
            pass

    def _orderdetail_recalc_total(self):
        try:
            gia = float(self.ui_order_detail.txtGiaBan.text().strip().replace(",", "") or 0)
        except ValueError:
            gia = 0
        so_luong = self.ui_order_detail.spinSoLuong.value()
        self.ui_order_detail.txtThanhTien.setText(f"{gia * so_luong:g}")

    def refresh_dependents(self):
        """Đồng bộ mọi màn hình liên quan sau mỗi lần Lưu/Xóa: bảng, combobox, thẻ số liệu, biểu đồ."""
        self.load_all_data_from_database()
        self.populate_all_comboboxes()
        self.load_statistics_data()
        self.update_all_charts()

    def init_sub_pages(self):
        self.page_product = QWidget()
        self.ui_product = Ui_ProductWidget()
        self.ui_product.setupUi(self.page_product)
        self.ui.stackedWidget.addWidget(self.page_product)

        self.page_comment = QWidget()
        self.ui_comment = Ui_CommentWidget()
        self.ui_comment.setupUi(self.page_comment)
        self.ui.stackedWidget.addWidget(self.page_comment)

        self.page_order = QWidget()
        self.ui_order = Ui_OrderWidget()
        self.ui_order.setupUi(self.page_order)
        self.ui.stackedWidget.addWidget(self.page_order)

        self.page_voucher = QWidget()
        self.ui_voucher = Ui_VoucherWidget()
        self.ui_voucher.setupUi(self.page_voucher)
        self.ui.stackedWidget.addWidget(self.page_voucher)

        self.page_livestream_detail = QWidget()
        self.ui_livestream_detail = Ui_LivestreamDetailWidget()
        self.ui_livestream_detail.setupUi(self.page_livestream_detail)
        self.ui.stackedWidget.addWidget(self.page_livestream_detail)

        self.page_order_detail = QWidget()
        self.ui_order_detail = Ui_OrderDetailWidget()
        self.ui_order_detail.setupUi(self.page_order_detail)
        self.ui.stackedWidget.addWidget(self.page_order_detail)

        self.page_payment = QWidget()
        self.ui_payment = Ui_PaymentWidget()
        self.ui_payment.setupUi(self.page_payment)
        self.ui.stackedWidget.addWidget(self.page_payment)

        self.page_statistics = QWidget()
        self.ui_statistics = Ui_StatisticsWidget()
        self.ui_statistics.setupUi(self.page_statistics)
        self.ui.stackedWidget.addWidget(self.page_statistics)

        self.page_livestream = QWidget()
        self.ui_livestream = Ui_LivestreamWidget()
        self.ui_livestream.setupUi(self.page_livestream)
        self.ui.stackedWidget.addWidget(self.page_livestream)

        self.page_seller = QWidget()
        self.ui_seller = Ui_SellerWidget()
        self.ui_seller.setupUi(self.page_seller)
        self.ui.stackedWidget.addWidget(self.page_seller)

        self.page_customer = QWidget()
        self.ui_customer = Ui_CustomerWidget()
        self.ui_customer.setupUi(self.page_customer)
        self.ui.stackedWidget.addWidget(self.page_customer)

    def connect_sidebar_buttons(self):
        self.ui.btnMenuProduct.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.page_product))
        self.ui.btnMenuComment.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.page_comment))
        self.ui.btnMenuOrder.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.page_order))
        self.ui.btnMenuVoucher.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.page_voucher))
        self.ui.btnMenuLivestreamDetail.clicked.connect(
            lambda: self.ui.stackedWidget.setCurrentWidget(self.page_livestream_detail))
        self.ui.btnMenuOrderDetail.clicked.connect(
            lambda: self.ui.stackedWidget.setCurrentWidget(self.page_order_detail))
        self.ui.btnMenuPayment.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.page_payment))
        self.ui.btnMenuStatistics.clicked.connect(self._open_statistics_tab)
        self.ui.btnMenuLivestream.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.page_livestream))
        self.ui.btnMenuSeller.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.page_seller))

        if hasattr(self.ui, "btnMenuCustomer"):
            self.ui.btnMenuCustomer.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.page_customer))
        if hasattr(self.ui, "btnMenuDashboard"):
            self.ui.btnMenuDashboard.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(0))
        if hasattr(self.ui, 'btnMenuLogout'):
            self.ui.btnMenuLogout.clicked.connect(self.process_logout)

    def process_logout(self):
        if self.login_window:
            self.login_window.show()
        self.close()

    # ==================== ĐẨY DỮ LIỆU THỰC TẾ TỪ SQLITE LÊN UI ====================
    def load_all_data_from_database(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # --- 1. THẺ THỐNG KÊ (STATS CARDS) ON MAIN DASHBOARD ---
            cursor.execute("SELECT COUNT(*) FROM KHACH_HANG")
            self.ui.lblCardCustomerNum.setText(str(cursor.fetchone()[0]))

            cursor.execute("SELECT COUNT(*) FROM LIVESTREAM")
            self.ui.lblCardLivestreamNum.setText(str(cursor.fetchone()[0]))

            cursor.execute("SELECT COUNT(*) FROM SAN_PHAM")
            self.ui.lblCardProductNum.setText(str(cursor.fetchone()[0]))

            cursor.execute("SELECT COUNT(*) FROM DON_HANG")
            self.ui.lblCardOrderNum.setText(str(cursor.fetchone()[0]))

            cursor.execute(
                "SELECT SUM(TongTien) FROM HOA_DON WHERE TrangThaiHD = 'Đã thanh toán' OR TrangThaiHD LIKE '%Da%'")
            total_revenue = cursor.fetchone()[0] or 0
            if total_revenue >= 1_000_000:
                self.ui.lblCardRevenueNum.setText(f"{total_revenue / 1_000_000:.1f}M")
            else:
                self.ui.lblCardRevenueNum.setText(f"{total_revenue:,.0f} VND")

            cursor.execute("SELECT COUNT(*) FROM VOUCHER WHERE TrangThai = 'Đang áp dụng' OR TrangThai LIKE '%ap%'")
            self.ui.lblCardVoucherNum.setText(str(cursor.fetchone()[0]))

            # --- 2. CÁC BẢNG TRÊN TRANG CHỦ DASHBOARD ---
            # Sản phẩm bán chạy: nhiều đơn "Đã giao" + hóa đơn "Đã thanh toán", cảnh báo tồn thấp
            cursor.execute("""
                SELECT p.MaSP, p.TenSP,
                       COALESCE(SUM(c.SoLuong), 0) as DaBan,
                       p.SoLuongTon
                FROM SAN_PHAM p
                JOIN CHI_TIET_DON_HANG c ON p.MaSP = c.MaSP
                JOIN DON_HANG dh ON c.MaDonHang = dh.MaDonHang AND dh.TrangThaiDH = 'Đã giao'
                JOIN HOA_DON hd ON hd.MaDonHang = dh.MaDonHang AND hd.TrangThaiHD = 'Đã thanh toán'
                GROUP BY p.MaSP ORDER BY DaBan DESC LIMIT 5
            """)
            top_rows = cursor.fetchall()
            self.populate_table(self.ui.tblTopProducts, top_rows)
            for c_idx, header in enumerate(["Mã SP", "Tên SP", "Đã bán", "Tồn kho"]):
                self.ui.tblTopProducts.setHorizontalHeaderItem(c_idx, QTableWidgetItem(header))
            for r_idx, row in enumerate(top_rows):
                if row[3] is not None and row[3] < 10:
                    item = self.ui.tblTopProducts.item(r_idx, 3)
                    if item:
                        item.setForeground(QtGui.QBrush(QtGui.QColor("#dc2626")))
                        item.setText(f"{row[3]} (sắp hết)")

            # Livestream đang chạy hoặc gần nhất (đồng bộ với tab Livestream)
            cursor.execute("""
                SELECT MaLive, TenLive, MaNguoiBan, TrangThai FROM LIVESTREAM
                ORDER BY CASE WHEN TrangThai LIKE '%ang%' THEN 0 ELSE 1 END, ThoiGianBatDau DESC
                LIMIT 5
            """)
            self.populate_table(self.ui.tblActiveLivestreams, cursor.fetchall())

            # --- 3. ĐỔ DỮ LIỆU VÀO CÁC PHÂN HỆ CON ---
            # Phân hệ Sản phẩm
            if hasattr(self.ui_product, 'tblProduct'):
                cursor.execute("SELECT MaSP, TenSP, GiaBan, SoLuongTon, HinhAnh, CASE WHEN SoLuongTon > 0 THEN 'Còn hàng' ELSE 'Hết hàng' END FROM SAN_PHAM")
                self.populate_table(self.ui_product.tblProduct, cursor.fetchall())

            # Phân hệ Bình luận
            if hasattr(self.ui_comment, 'tblComment'):
                cursor.execute("SELECT MaBinhLuan, MaLive, NguoiBinhLuan, NoiDung, ThoiGian FROM BINH_LUAN")
                self.populate_table(self.ui_comment.tblComment, cursor.fetchall())

            # Phân hệ Đơn hàng (query dùng chung với refresh_order_table)
            self.refresh_order_table()

            # Phân hệ Khách hàng (Đã sửa từ tblCustomer sang tblKhachHang)
            if hasattr(self.ui_customer, 'tblKhachHang'):
                cursor.execute("SELECT MaKhachHang, HoTen, SoDienThoai, Email, DiaChi, '2026-07-15' FROM KHACH_HANG")
                self.populate_table(self.ui_customer.tblKhachHang, cursor.fetchall())

            # Phân hệ Người bán
            if hasattr(self.ui_seller, 'tblSeller'):
                cursor.execute("SELECT MaNguoiBan, HoTen, SoDienThoai, Email, TenCuaHang FROM NGUOI_BAN")
                self.populate_table(self.ui_seller.tblSeller, cursor.fetchall())

            # Phân hệ Voucher
            if hasattr(self.ui_voucher, 'tblVoucher'):
                cursor.execute("""
                    SELECT
                        MaVoucher,
                        COALESCE(LoaiUuDai, 'Số tiền cố định (VND)') AS LoaiUuDai,
                        GiaTriGiam,
                        DieuKienApDung,
                        COALESCE(GiamToiDa, GiaTriGiam) AS GiamToiDa,
                        NgayBatDau,
                        NgayKetThuc,
                        TrangThai
                    FROM VOUCHER
                """)
                self.populate_table(self.ui_voucher.tblVoucher, cursor.fetchall())

            # Phân hệ Livestream
            if hasattr(self.ui_livestream, 'tblLivestream'):
                cursor.execute("SELECT MaLive, TenLive, ThoiGianBatDau, MaNguoiBan, TrangThai FROM LIVESTREAM")
                self.populate_table(self.ui_livestream.tblLivestream, cursor.fetchall())

            # Phân hệ Livestream Chi Tiết
            if hasattr(self.ui_livestream_detail, 'tblLivestreamDetail'):
                cursor.execute("""
                    SELECT 
                        (lsp.MaLive || '-' || lsp.MaSP) AS MaChiTietLive,
                        l.TenLive,
                        lsp.MaSP,
                        p.TenSP,
                        p.SoLuongTon,
                        p.GiaBan
                    FROM LIVESTREAM_SAN_PHAM lsp
                    JOIN LIVESTREAM l ON lsp.MaLive = l.MaLive
                    JOIN SAN_PHAM p ON lsp.MaSP = p.MaSP
                """)
                self.populate_table(self.ui_livestream_detail.tblLivestreamDetail, cursor.fetchall())

            # Phân hệ Chi Tiết Đơn Hàng (query dùng chung với refresh_order_detail_table)
            self.refresh_order_detail_table()

            # Phân hệ Thanh Toán (Payment)
            if hasattr(self.ui_payment, 'tblPayment'):
                cursor.execute("SELECT MaHoaDon, MaDonHang, ThoiGianLap, TongTien, PhuongThucTT, TrangThaiHD FROM HOA_DON")
                self.populate_table(self.ui_payment.tblPayment, cursor.fetchall())

            # Phân hệ Thống Kê (Statistics)
            if hasattr(self.ui_statistics, 'tblStatistics'):
                cursor.execute("""
                    SELECT 
                        l.MaLive,
                        l.TenLive,
                        nb.HoTen,
                        COALESCE((
                            SELECT SUM(dh.TongTien) 
                            FROM DON_HANG dh
                            JOIN BINH_LUAN bl ON dh.MaBinhLuan = bl.MaBinhLuan
                            WHERE bl.MaLive = l.MaLive
                        ), 0) AS DoanhThu,
                        COALESCE((
                            SELECT COUNT(dh.MaDonHang) 
                            FROM DON_HANG dh
                            JOIN BINH_LUAN bl ON dh.MaBinhLuan = bl.MaBinhLuan
                            WHERE bl.MaLive = l.MaLive
                        ), 0) AS SoDonDat,
                        COALESCE((
                            SELECT SUM(ct.SoLuong) 
                            FROM CHI_TIET_DON_HANG ct
                            JOIN DON_HANG dh ON ct.MaDonHang = dh.MaDonHang
                            JOIN BINH_LUAN bl ON dh.MaBinhLuan = bl.MaBinhLuan
                            WHERE bl.MaLive = l.MaLive
                        ), 0) AS SanPhamDaBan
                    FROM LIVESTREAM l
                    LEFT JOIN NGUOI_BAN nb ON l.MaNguoiBan = nb.MaNguoiBan
                """)
                self.populate_table(self.ui_statistics.tblStatistics, cursor.fetchall())
                self._format_stats_table()

                # Cập nhật Thẻ Thống kê (Summary Cards) trên trang Thống kê
                cursor.execute("SELECT SUM(TongTien) FROM HOA_DON WHERE TrangThaiHD = 'Đã thanh toán' OR TrangThaiHD LIKE '%Da%'")
                tot_rev = cursor.fetchone()[0] or 0
                cursor.execute("SELECT COUNT(*) FROM DON_HANG")
                tot_ord = cursor.fetchone()[0] or 0
                avg_val = tot_rev / tot_ord if tot_ord > 0 else 0

                if hasattr(self.ui_statistics, 'lblTotalRevenue'):
                    self.ui_statistics.lblTotalRevenue.setText(f"{tot_rev:,.0f} VNĐ")
                if hasattr(self.ui_statistics, 'lblTotalOrders'):
                    self.ui_statistics.lblTotalOrders.setText(f"{tot_ord:,} đơn đặt")
                if hasattr(self.ui_statistics, 'lblAvgOrderValue'):
                    self.ui_statistics.lblAvgOrderValue.setText(f"{avg_val:,.0f} VNĐ")

            conn.close()
        except Exception as e:
            print(f"Lỗi truy vấn SQL: {e}")

    def populate_table(self, table_widget, data):
        table_widget.setRowCount(0)
        for row_idx, row_data in enumerate(data):
            table_widget.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                val_str = str(value if value is not None else "")
                item = QTableWidgetItem(val_str)
                # Stylesheet nền sáng nhưng màu chữ mặc định theo theme bị trùng nền — ép chữ đen
                item.setForeground(QtGui.QBrush(QtGui.QColor("#000000")))
                table_widget.setItem(row_idx, col_idx, item)
        if hasattr(table_widget, 'horizontalHeader') and table_widget.horizontalHeader():
            table_widget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)

    # ==================== PHÂN HỆ QUẢN LÝ NGƯỜI BÁN & SẢN PHẨM KHỚP SQLITE ====================

    def setup_seller_product_actions(self):
        # Biến lưu trữ ID dòng đang chọn
        self.selected_product_id = None
        self.selected_seller_id = None

        # Thiết lập chế độ chọn dòng và không cho sửa trực tiếp trên bảng
        if hasattr(self.ui_product, 'tblProduct'):
            self.ui_product.tblProduct.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
            self.ui_product.tblProduct.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
            self.ui_product.tblProduct.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
            self.ui_product.tblProduct.cellClicked.connect(self.product_cell_clicked)
            self.ui_product.btnThem.clicked.connect(self.product_action_them)
            self.ui_product.btnSua.clicked.connect(self._on_product_edit_button)
            self.ui_product.btnLuu.clicked.connect(self.product_action_save)
            self.ui_product.btnLuu.setVisible(False)
            self._product_action = None
            self._set_product_form_enabled(False)
            self.ui_product.btnXoa.clicked.connect(self.product_action_xoa)
            self.ui_product.btnHuy.clicked.connect(self.product_action_clear)
            self.ui_product.btnXuatExcel.clicked.connect(self.product_export_excel)
            self.ui_product.btnSearch.clicked.connect(self.product_search)
            self.ui_product.btnRefresh.clicked.connect(self.product_refresh)
            if hasattr(self.ui_product, 'btnSelectImage'):
                self.ui_product.btnSelectImage.clicked.connect(self.product_select_image)

            # Đồng bộ combobox trạng thái và spinbox tồn kho
            self.ui_product.spinSoLuongTon.valueChanged.connect(self.product_spin_changed)
            self.ui_product.cbTrangThai.currentIndexChanged.connect(self.product_status_changed)

        if hasattr(self.ui_seller, 'tblSeller'):
            self.ui_seller.tblSeller.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
            self.ui_seller.tblSeller.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
            self.ui_seller.tblSeller.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
            self.ui_seller.tblSeller.cellClicked.connect(self.seller_cell_clicked)
            self.ui_seller.btnThem.clicked.connect(self.seller_action_them)
            self.ui_seller.btnSua.clicked.connect(self._on_seller_edit_button)
            self.ui_seller.btnLuu.clicked.connect(self.seller_action_save)
            self.ui_seller.btnLuu.setVisible(False)
            self._seller_action = None
            self._set_seller_form_enabled(False)
            self.ui_seller.btnXoa.clicked.connect(self.seller_action_xoa)
            self.ui_seller.btnHuy.clicked.connect(self.seller_action_clear)
            self.ui_seller.btnXuatExcel.clicked.connect(self.seller_export_excel)
            self.ui_seller.btnSearch.clicked.connect(self.seller_search)
            self.ui_seller.btnRefresh.clicked.connect(self.seller_refresh)

    def setup_all_other_actions(self):
        # Initial populate of combo boxes
        self.populate_all_comboboxes()

        # Comment Tab
        comment_mappings = [
            ("MaLive", self.ui_comment.cbLivestream, "combo_text"),
            ("NguoiBinhLuan", self.ui_comment.txtTenNguoiDung, "text"),
            ("NoiDung", self.ui_comment.txtNoiDung, "text"),
            ("ThoiGian", self.ui_comment.dateThoiGian, "datetime")
        ]
        self.comment_controller = DbFormController(
            self, self.ui_comment, self.ui_comment.tblComment, "BINH_LUAN", "MaBinhLuan", comment_mappings, self.refresh_comment_table
        )

        # Customer Tab
        customer_mappings = [
            ("HoTen", self.ui_customer.txtTenKhachHang, "text"),
            ("SoDienThoai", self.ui_customer.txtSoDienThoai, "text"),
            ("Email", self.ui_customer.txtEmail, "text"),
            ("DiaChi", self.ui_customer.txtDiaChi, "text")
        ]
        self.customer_controller = DbFormController(
            self, self.ui_customer, self.ui_customer.tblKhachHang, "KHACH_HANG", "MaKhachHang", customer_mappings, self.refresh_customer_table
        )

        # Livestream Tab
        livestream_mappings = [
            ("TenLive", self.ui_livestream.txtTieuDe, "text"),
            ("MaNguoiBan", self.ui_livestream.cbNguoiBan, "combo_text"),
            ("ThoiGianBatDau", self.ui_livestream.dateNgayKhaiMac, "datetime"),
            ("TrangThai", self.ui_livestream.cbTrangThai, "combo_text")
        ]
        self.livestream_controller = DbFormController(
            self, self.ui_livestream, self.ui_livestream.tblLivestream, "LIVESTREAM", "MaLive", livestream_mappings, self.refresh_livestream_table
        )

        # Livestream Detail Tab
        livestream_detail_mappings = [
            ("MaLive", self.ui_livestream_detail.cbLivestream, "combo_text"),
            ("MaSP", self.ui_livestream_detail.cbSanPham, "combo_text"),
            ("SoLuongTon", self.ui_livestream_detail.spinSoLuongGioiThieu, "spin"),
            ("GiaBan", self.ui_livestream_detail.txtGiaKhuyenMai, "text")
        ]
        self.livestream_detail_controller = DbFormController(
            self, self.ui_livestream_detail, self.ui_livestream_detail.tblLivestreamDetail, "LIVESTREAM_SAN_PHAM", "MaChiTietLive", livestream_detail_mappings, self.refresh_livestream_detail_table
        )

        # Order Tab
        order_mappings = [
            ("NgayDat", self.ui_order.dateNgayDat, "datetime"),
            ("MaLive", self.ui_order.cbLivestream, "combo_text"),
            ("HoTen", self.ui_order.txtTenKhachHang, "text"),
            ("SoDienThoai", self.ui_order.txtSoDienThoai, "text"),
            ("DiaChi", self.ui_order.txtDiaChiGiao, "text"),
            ("MaVoucher", self.ui_order.cbVoucher, "combo_text"),
            ("TrangThaiDH", self.ui_order.cbTrangThai, "combo_text"),
            ("TongTien", self.ui_order.txtTongTien, "text")
        ]
        self.order_controller = DbFormController(
            self, self.ui_order, self.ui_order.tblOrder, "DON_HANG", "MaDonHang", order_mappings, self.refresh_order_table
        )

        # Order Detail Tab
        order_detail_mappings = [
            ("MaDonHang", self.ui_order_detail.cbDonHang, "combo_text"),
            ("MaSP", self.ui_order_detail.cbSanPham, "combo_text"),
            ("SoLuong", self.ui_order_detail.spinSoLuong, "spin"),
            ("DonGia", self.ui_order_detail.txtGiaBan, "text"),
            ("ThanhTien", self.ui_order_detail.txtThanhTien, "text")
        ]
        self.order_detail_controller = DbFormController(
            self, self.ui_order_detail, self.ui_order_detail.tblOrderDetail, "CHI_TIET_DON_HANG", "MaChiTietDH", order_detail_mappings, self.refresh_order_detail_table
        )

        # Payment Tab
        payment_mappings = [
            ("MaDonHang", self.ui_payment.cbDonHang, "combo_text"),
            ("ThoiGianLap", self.ui_payment.dateThanhToan, "datetime"),
            ("TongTien", self.ui_payment.txtSoTien, "text"),
            ("PhuongThucTT", self.ui_payment.cbPhuongThuc, "combo_text"),
            ("TrangThaiHD", self.ui_payment.cbTrangThai, "combo_text")
        ]
        self.payment_controller = DbFormController(
            self, self.ui_payment, self.ui_payment.tblPayment, "HOA_DON", "MaHoaDon", payment_mappings, self.refresh_payment_table
        )

        # Voucher Tab
        voucher_mappings = [
            ("MaVoucher", self.ui_voucher.txtMaVoucher, "text"),
            ("LoaiUuDai", self.ui_voucher.cbLoaiVoucher, "combo_text"),
            ("GiaTriGiam", self.ui_voucher.txtGiaTriGiam, "text"),
            ("DieuKienApDung", self.ui_voucher.txtDonHangToiThieu, "text"),
            ("GiamToiDa", self.ui_voucher.txtGiamToiDa, "text"),
            ("NgayBatDau", self.ui_voucher.dateStart, "date"),
            ("NgayKetThuc", self.ui_voucher.dateEnd, "date"),
            ("TrangThai", self.ui_voucher.cbKichHoat, "combo_text")
        ]
        self.voucher_controller = DbFormController(
            self, self.ui_voucher, self.ui_voucher.tblVoucher, "VOUCHER", "MaVoucher", voucher_mappings, self.refresh_voucher_table
        )

        # Statistics Tab
        self.setup_statistics_actions()

    def populate_all_comboboxes(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # 1. Fetch Livestreams
            cursor.execute("SELECT MaLive, TenLive FROM LIVESTREAM")
            lives = cursor.fetchall()
            live_items = [f"{row[0]} - {row[1]}" for row in lives]
            
            self.ui_comment.cbLivestream.clear()
            self.ui_comment.cbLivestream.addItems(live_items)
            
            self.ui_livestream_detail.cbLivestream.clear()
            self.ui_livestream_detail.cbLivestream.addItems(live_items)
            
            self.ui_order.cbLivestream.clear()
            self.ui_order.cbLivestream.addItems(live_items)
            
            # 2. Fetch Sellers
            cursor.execute("SELECT MaNguoiBan, HoTen FROM NGUOI_BAN")
            sellers = cursor.fetchall()
            seller_items = [f"{row[0]} - {row[1]}" for row in sellers]
            self.ui_livestream.cbNguoiBan.clear()
            self.ui_livestream.cbNguoiBan.addItems(seller_items)
            
            # 3. Fetch Products
            cursor.execute("SELECT MaSP, TenSP FROM SAN_PHAM")
            prods = cursor.fetchall()
            prod_items = [f"{row[0]} - {row[1]}" for row in prods]
            
            self.ui_livestream_detail.cbSanPham.clear()
            self.ui_livestream_detail.cbSanPham.addItems(prod_items)
            
            self.ui_order_detail.cbSanPham.clear()
            self.ui_order_detail.cbSanPham.addItems(prod_items)
            
            # 4. Fetch Vouchers
            cursor.execute("SELECT MaVoucher, TenVoucher FROM VOUCHER")
            vouchers = cursor.fetchall()
            voucher_items = ["Không"] + [f"{row[0]} - {row[1]}" for row in vouchers]
            self.ui_order.cbVoucher.clear()
            self.ui_order.cbVoucher.addItems(voucher_items)
            
            # 5. Fetch Orders
            cursor.execute("SELECT MaDonHang FROM DON_HANG")
            orders = cursor.fetchall()
            order_items = [row[0] for row in orders]
            
            self.ui_order_detail.cbDonHang.clear()
            self.ui_order_detail.cbDonHang.addItems(order_items)
            
            self.ui_payment.cbDonHang.clear()
            self.ui_payment.cbDonHang.addItems(order_items)
            
            conn.close()
        except Exception as e:
            print(f"Error populating comboboxes: {e}")

    def refresh_comment_table(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT MaBinhLuan, MaLive, NguoiBinhLuan, NoiDung, ThoiGian FROM BINH_LUAN")
            self.populate_table(self.ui_comment.tblComment, cursor.fetchall())
            conn.close()
            self.populate_all_comboboxes()
        except Exception as e:
            print(f"Error refreshing comments: {e}")

    def refresh_customer_table(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT MaKhachHang, HoTen, SoDienThoai, Email, DiaChi, '2026-07-15' FROM KHACH_HANG")
            self.populate_table(self.ui_customer.tblKhachHang, cursor.fetchall())
            conn.close()
            self.populate_all_comboboxes()
        except Exception as e:
            print(f"Error refreshing customers: {e}")

    def refresh_livestream_table(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT MaLive, TenLive, ThoiGianBatDau, MaNguoiBan, TrangThai FROM LIVESTREAM")
            self.populate_table(self.ui_livestream.tblLivestream, cursor.fetchall())
            conn.close()
            self.populate_all_comboboxes()
        except Exception as e:
            print(f"Error refreshing livestreams: {e}")

    def refresh_livestream_detail_table(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    (lsp.MaLive || '-' || lsp.MaSP) AS MaChiTietLive,
                    l.TenLive,
                    lsp.MaSP,
                    p.TenSP,
                    p.SoLuongTon,
                    p.GiaBan
                FROM LIVESTREAM_SAN_PHAM lsp
                JOIN LIVESTREAM l ON lsp.MaLive = l.MaLive
                JOIN SAN_PHAM p ON lsp.MaSP = p.MaSP
            """)
            self.populate_table(self.ui_livestream_detail.tblLivestreamDetail, cursor.fetchall())
            conn.close()
            self.populate_all_comboboxes()
        except Exception as e:
            print(f"Error refreshing livestream details: {e}")

    def refresh_order_table(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    dh.MaDonHang,
                    dh.NgayDat,
                    kh.HoTen,
                    kh.SoDienThoai,
                    dh.TongTien,
                    COALESCE(v.TenVoucher, dh.MaVoucher),
                    dh.TrangThaiDH,
                    COALESCE((SELECT GROUP_CONCAT(p.TenSP, ', ')
                              FROM CHI_TIET_DON_HANG ct JOIN SAN_PHAM p ON ct.MaSP = p.MaSP
                              WHERE ct.MaDonHang = dh.MaDonHang), '(chưa có SP)')
                FROM DON_HANG dh
                LEFT JOIN KHACH_HANG kh ON dh.MaKhachHang = kh.MaKhachHang
                LEFT JOIN VOUCHER v ON dh.MaVoucher = v.MaVoucher
            """)
            self.populate_table(self.ui_order.tblOrder, cursor.fetchall())
            conn.close()
            self.populate_all_comboboxes()
        except Exception as e:
            print(f"Error refreshing orders: {e}")

    def refresh_order_detail_table(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    (ct.MaDonHang || '-' || ct.MaSP) AS MaChiTietDH,
                    ct.MaDonHang,
                    ct.MaSP,
                    p.TenSP,
                    ct.SoLuong,
                    ct.DonGia,
                    ct.ThanhTien,
                    COALESCE(v.TenVoucher, dh.MaVoucher, 'Không')
                FROM CHI_TIET_DON_HANG ct
                JOIN SAN_PHAM p ON ct.MaSP = p.MaSP
                LEFT JOIN DON_HANG dh ON ct.MaDonHang = dh.MaDonHang
                LEFT JOIN VOUCHER v ON dh.MaVoucher = v.MaVoucher
            """)
            self.populate_table(self.ui_order_detail.tblOrderDetail, cursor.fetchall())
            conn.close()
            self.populate_all_comboboxes()
        except Exception as e:
            print(f"Error refreshing order details: {e}")

    def refresh_payment_table(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT MaHoaDon, MaDonHang, ThoiGianLap, TongTien, PhuongThucTT, TrangThaiHD FROM HOA_DON")
            self.populate_table(self.ui_payment.tblPayment, cursor.fetchall())
            conn.close()
            self.populate_all_comboboxes()
        except Exception as e:
            print(f"Error refreshing payments: {e}")

    def refresh_voucher_table(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    MaVoucher,
                    COALESCE(LoaiUuDai, 'Số tiền cố định (VND)') AS LoaiUuDai,
                    GiaTriGiam,
                    DieuKienApDung,
                    COALESCE(GiamToiDa, GiaTriGiam) AS GiamToiDa,
                    NgayBatDau,
                    NgayKetThuc,
                    TrangThai
                FROM VOUCHER
            """)
            self.populate_table(self.ui_voucher.tblVoucher, cursor.fetchall())
            conn.close()
            self.populate_all_comboboxes()
        except Exception as e:
            print(f"Error refreshing vouchers: {e}")

    def setup_statistics_actions(self):
        if hasattr(self.ui_statistics, 'btnFilter'):
            self.ui_statistics.btnFilter.clicked.connect(self.statistics_filter)
        if hasattr(self.ui_statistics, 'btnRefresh'):
            self.ui_statistics.btnRefresh.clicked.connect(self.statistics_refresh)
        if hasattr(self.ui_statistics, 'btnXuatExcel'):
            self.ui_statistics.btnXuatExcel.clicked.connect(self.statistics_export_excel)
        
        # Mặc định bao trọn toàn bộ dữ liệu (từ trước đến nay)
        self.ui_statistics.dateStart.setDate(QtCore.QDate(2020, 1, 1))
        self.ui_statistics.dateEnd.setDate(QtCore.QDate.currentDate().addDays(1))
        # Nạp báo cáo toàn thời gian ngay khi khởi động
        self.load_statistics_data()

    def load_statistics_data(self, start_date=None, end_date=None):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            query = """
                SELECT 
                    l.MaLive,
                    l.TenLive,
                    nb.HoTen,
                    COALESCE((
                        SELECT SUM(dh.TongTien) 
                        FROM DON_HANG dh
                        JOIN BINH_LUAN bl ON dh.MaBinhLuan = bl.MaBinhLuan
                        WHERE bl.MaLive = l.MaLive
                    ), 0) AS DoanhThu,
                    COALESCE((
                        SELECT COUNT(dh.MaDonHang) 
                        FROM DON_HANG dh
                        JOIN BINH_LUAN bl ON dh.MaBinhLuan = bl.MaBinhLuan
                        WHERE bl.MaLive = l.MaLive
                    ), 0) AS SoDonDat,
                    COALESCE((
                        SELECT SUM(ct.SoLuong) 
                        FROM CHI_TIET_DON_HANG ct
                        JOIN DON_HANG dh ON ct.MaDonHang = dh.MaDonHang
                        JOIN BINH_LUAN bl ON dh.MaBinhLuan = bl.MaBinhLuan
                        WHERE bl.MaLive = l.MaLive
                    ), 0) AS SanPhamDaBan
                FROM LIVESTREAM l
                LEFT JOIN NGUOI_BAN nb ON l.MaNguoiBan = nb.MaNguoiBan
            """
            
            params = []
            if start_date and end_date:
                query += " WHERE l.ThoiGianBatDau BETWEEN ? AND ?"
                params = [start_date + " 00:00:00", end_date + " 23:59:59"]
                
            cursor.execute(query, params)
            rows = cursor.fetchall()
            self.populate_table(self.ui_statistics.tblStatistics, rows)
            self._format_stats_table()

            # Vẽ lại biểu đồ thống kê theo đúng bộ dữ liệu (kể cả khi lọc ngày)
            self.update_statistics_charts(rows, start_date, end_date)
            
            # Recalculate summary cards
            tot_rev = sum(row[3] for row in rows)
            tot_ord = sum(row[4] for row in rows)
            avg_val = tot_rev / tot_ord if tot_ord > 0 else 0
            
            if hasattr(self.ui_statistics, 'lblTotalRevenue'):
                self.ui_statistics.lblTotalRevenue.setText(f"{tot_rev:,.0f} VNĐ")
            if hasattr(self.ui_statistics, 'lblTotalOrders'):
                self.ui_statistics.lblTotalOrders.setText(f"{tot_ord:,} đơn đặt")
            if hasattr(self.ui_statistics, 'lblAvgOrderValue'):
                self.ui_statistics.lblAvgOrderValue.setText(f"{avg_val:,.0f} VNĐ")
                
            conn.close()
        except Exception as e:
            print(f"Error loading stats data: {e}")

    def statistics_filter(self):
        start_date = self.ui_statistics.dateStart.date().toString("yyyy-MM-dd")
        end_date = self.ui_statistics.dateEnd.date().toString("yyyy-MM-dd")
        self.load_statistics_data(start_date, end_date)
        if hasattr(self.ui_statistics, 'lblStatus'):
            self.ui_statistics.lblStatus.setText(f"Trạng thái: Đã lọc từ {start_date} đến {end_date}")

    def _open_statistics_tab(self):
        # Mở tab Thống kê: chuyển trang + tự nạp báo cáo toàn thời gian
        self.ui.stackedWidget.setCurrentWidget(self.page_statistics)
        self.load_statistics_data()

    def statistics_refresh(self):
        # Làm mới = xem lại toàn bộ dữ liệu từ trước đến nay
        self.ui_statistics.dateStart.setDate(QtCore.QDate(2020, 1, 1))
        self.ui_statistics.dateEnd.setDate(QtCore.QDate.currentDate().addDays(1))
        self.load_statistics_data()
        if hasattr(self.ui_statistics, 'lblStatus'):
            self.ui_statistics.lblStatus.setText("Trạng thái: Đã làm mới báo cáo.")

    def statistics_export_excel(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Lưu file Excel thống kê", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if not path:
            return
            
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "ThongKeDoanhThu"
            
            headers = ["Mã Stream", "Tiêu Đề Livestream", "Người Bán", "Doanh Thu (VND)", "Số Đơn Đặt", "Sản Phẩm Đã Bán"]
            ws.append(headers)
            
            for row in range(self.ui_statistics.tblStatistics.rowCount()):
                row_data = []
                for col in range(self.ui_statistics.tblStatistics.columnCount()):
                    item = self.ui_statistics.tblStatistics.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
                
            wb.save(path)
            QMessageBox.information(self, "Thành công", f"Đã xuất file báo cáo Excel thành công tại:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể lưu file Excel: {e}")

    # --- BIỂU ĐỒ QTCHARTS (Dashboard + Statistics) ---
    def setup_charts_area(self):
        """Chuẩn bị QChartView gắn vào các frame có sẵn; thiếu PyQt6-Charts thì để nguyên placeholder."""
        self._chart_views = {}
        if not HAS_QTCHARTS:
            for lbl in (getattr(self.ui, 'lblChart1Placeholder', None),
                        getattr(self.ui, 'lblChart2Placeholder', None),
                        getattr(self.ui_statistics, 'lblChartPlaceholder', None)):
                if lbl:
                    lbl.setText("Cài PyQt6-Charts để xem biểu đồ: pip install PyQt6-Charts")
            return

        def make_view(layout, placeholder, max_height=None):
            if placeholder:
                placeholder.hide()
            view = QChartView()
            view.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
            view.setMinimumHeight(180)
            if max_height:
                view.setMaximumHeight(max_height)
            layout.addWidget(view)
            return view

        # Dashboard: bar doanh thu + pie số đơn — giới hạn cao để cân đối với 2 bảng bên dưới
        self._chart_views['dash_revenue'] = make_view(self.ui.verticalLayout_chart1, self.ui.lblChart1Placeholder, max_height=260)
        self._chart_views['dash_orders'] = make_view(self.ui.verticalLayout_chart2, self.ui.lblChart2Placeholder, max_height=260)

        # Statistics: bar doanh thu theo livestream + pie doanh thu theo sản phẩm (đặt cạnh nhau)
        stats_row = QtWidgets.QHBoxLayout()
        self.ui_statistics.verticalLayout_innerChart.addLayout(stats_row)
        self._chart_views['stats_bar'] = make_view(stats_row, self.ui_statistics.lblChartPlaceholder)
        self._chart_views['stats_pie'] = make_view(stats_row, None)

    def refactor_statistics_ui(self):
        """Tối ưu bố cục trang Thống kê: xếp dọc (KPI -> biểu đồ full-width -> bảng chi tiết),
        bảng canh số phải, xen màu, dòng thoáng. Chỉ chỉnh runtime, không đụng file .ui."""
        us = self.ui_statistics
        reports = us.layoutReports  # QHBox: [khối bảng | khối biểu đồ] -> chuyển sang dọc
        if reports.count() >= 2:
            reports.setDirection(QtWidgets.QBoxLayout.Direction.TopToBottom)
            reports.setSpacing(16)
            chart_item = reports.takeAt(1)   # khối biểu đồ đang ở cột phải
            reports.insertItem(0, chart_item)  # đưa lên trên, bảng xuống dưới -> đọc xu hướng trước

        # Biểu đồ full-width cần cao hơn để dễ đọc
        us.frameChart.setMinimumHeight(300)

        # Bảng chi tiết: chọn theo dòng, không sửa tay, xen màu, dòng thoáng
        tbl = us.tblStatistics
        tbl.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        tbl.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        tbl.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setAlternatingRowColors(True)
        tbl.setStyleSheet(tbl.styleSheet() + "\nQTableWidget { alternate-background-color: #faf5ff; }")
        tbl.verticalHeader().setVisible(False)
        tbl.verticalHeader().setDefaultSectionSize(38)
        hh = tbl.horizontalHeader()
        hh.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)  # cột Tiêu đề co giãn
        for c in (0, 2, 3, 4, 5):
            hh.setSectionResizeMode(c, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)

    def _format_stats_table(self):
        """Canh phải + phân tách nghìn cho các cột số của bảng thống kê."""
        tbl = self.ui_statistics.tblStatistics
        right = QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignVCenter
        for r in range(tbl.rowCount()):
            for c in (3, 4, 5):  # Doanh thu, Số đơn, Sản phẩm đã bán
                it = tbl.item(r, c)
                if not it:
                    continue
                try:
                    it.setText(f"{float(it.text().replace(',', '').strip()):,.0f}")
                except ValueError:
                    pass
                it.setTextAlignment(right)

    def _set_chart(self, key, chart):
        view = self._chart_views.get(key)
        if view:
            old = view.chart()
            view.setChart(chart)
            if old is not None:
                old.deleteLater()

    @staticmethod
    def _make_bar_chart(title, categories, values, series_label):
        if not values:
            categories, values = ["(chưa có dữ liệu)"], [0]
        values = [float(v or 0) for v in values]
        bar_set = QBarSet(series_label)
        for v in values:
            bar_set.append(v)
        series = QBarSeries()
        series.append(bar_set)
        chart = QChart()
        chart.addSeries(series)
        chart.setTitle(title)
        axis_x = QBarCategoryAxis()
        axis_x.append(categories)
        chart.addAxis(axis_x, QtCore.Qt.AlignmentFlag.AlignBottom)
        series.attachAxis(axis_x)
        axis_y = QValueAxis()
        axis_y.setRange(0, max(values) * 1.1 or 1)
        chart.addAxis(axis_y, QtCore.Qt.AlignmentFlag.AlignLeft)
        series.attachAxis(axis_y)
        chart.legend().setVisible(False)
        chart.setMargins(QtCore.QMargins(4, 4, 4, 4))
        chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        return chart

    @staticmethod
    def _make_pie_chart(title, pairs):
        series = QPieSeries()
        for label, value in pairs:
            if value:
                series.append(f"{label} ({value:g})", value)
        series.setLabelsVisible(True)
        chart = QChart()
        chart.addSeries(series)
        chart.setTitle(title)
        chart.legend().setVisible(True)
        chart.legend().setAlignment(QtCore.Qt.AlignmentFlag.AlignRight)
        chart.setMargins(QtCore.QMargins(4, 4, 4, 4))
        chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        return chart

    def update_all_charts(self):
        """Biểu đồ Dashboard: bar doanh thu theo tháng + pie số đơn theo trạng thái."""
        if not HAS_QTCHARTS:
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # Doanh thu theo tháng từ hóa đơn đã thanh toán
            cursor.execute("""
                SELECT strftime('%Y-%m', ThoiGianLap) AS Thang, SUM(TongTien)
                FROM HOA_DON WHERE TrangThaiHD = 'Đã thanh toán' AND ThoiGianLap IS NOT NULL
                GROUP BY Thang ORDER BY Thang
            """)
            rows = cursor.fetchall()
            self._set_chart('dash_revenue', self._make_bar_chart(
                "Doanh thu theo tháng (VND)", [r[0] or "?" for r in rows], [r[1] or 0 for r in rows], "Doanh thu"))

            # Số đơn theo trạng thái
            cursor.execute("SELECT TrangThaiDH, COUNT(*) FROM DON_HANG GROUP BY TrangThaiDH")
            self._set_chart('dash_orders', self._make_pie_chart(
                "Số đơn hàng theo trạng thái", cursor.fetchall()))
            conn.close()
        except Exception as e:
            print(f"Lỗi vẽ biểu đồ dashboard: {e}")

    def update_statistics_charts(self, stat_rows, start_date=None, end_date=None):
        """Biểu đồ Statistics: bar doanh thu theo livestream + pie doanh thu theo sản phẩm."""
        if not HAS_QTCHARTS:
            return
        try:
            suffix = f" ({start_date} → {end_date})" if start_date and end_date else ""
            self._set_chart('stats_bar', self._make_bar_chart(
                "Doanh thu theo livestream" + suffix,
                [r[0] for r in stat_rows], [r[3] or 0 for r in stat_rows], "Doanh thu"))

            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            params = []
            date_cond = ""
            if start_date and end_date:
                date_cond = " AND dh.NgayDat BETWEEN ? AND ?"
                params = [start_date + " 00:00:00", end_date + " 23:59:59"]
            cursor.execute(f"""
                SELECT p.TenSP, SUM(ct.ThanhTien)
                FROM CHI_TIET_DON_HANG ct
                JOIN SAN_PHAM p ON ct.MaSP = p.MaSP
                JOIN DON_HANG dh ON ct.MaDonHang = dh.MaDonHang
                WHERE dh.TrangThaiDH != 'Đã hủy'{date_cond}
                GROUP BY p.MaSP ORDER BY 2 DESC
            """, params)
            self._set_chart('stats_pie', self._make_pie_chart(
                "Tỷ trọng doanh thu theo sản phẩm" + suffix, cursor.fetchall()))
            conn.close()
        except Exception as e:
            print(f"Lỗi vẽ biểu đồ thống kê: {e}")

    # --- HÀM TRUY CẬP ẢNH SẢN PHẨM ---
    def load_product_image(self, image_path):
        if not image_path:
            self.ui_product.lblImagePreview.setText("Không có ảnh")
            self.ui_product.lblImagePreview.setPixmap(QtGui.QPixmap())
            return

        current_dir = os.path.dirname(os.path.abspath(__file__))
        parent_dir = os.path.dirname(current_dir)

        if not os.path.isabs(image_path):
            p1 = os.path.join(current_dir, image_path)
            p2 = os.path.join(parent_dir, image_path)
            p3 = os.path.join(current_dir, "Product", image_path)
            if os.path.exists(p1):
                image_path = p1
            elif os.path.exists(p2):
                image_path = p2
            elif os.path.exists(p3):
                image_path = p3

        if os.path.exists(image_path):
            pixmap = QtGui.QPixmap(image_path)
            if not pixmap.isNull():
                scaled = pixmap.scaled(
                    self.ui_product.lblImagePreview.width(),
                    self.ui_product.lblImagePreview.height(),
                    QtCore.Qt.AspectRatioMode.KeepAspectRatio,
                    QtCore.Qt.TransformationMode.SmoothTransformation
                )
                self.ui_product.lblImagePreview.setPixmap(scaled)
            else:
                self.ui_product.lblImagePreview.setText("Lỗi định dạng ảnh")
        else:
            self.ui_product.lblImagePreview.setText(f"Không tìm thấy: {os.path.basename(image_path)}")

    # --- CHI TIẾT SẢN PHẨM ---
    def load_product_data(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT MaSP, TenSP, GiaBan, SoLuongTon, HinhAnh, CASE WHEN SoLuongTon > 0 THEN 'Còn hàng' ELSE 'Hết hàng' END FROM SAN_PHAM")
            self.populate_table(self.ui_product.tblProduct, cursor.fetchall())
            conn.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể tải dữ liệu sản phẩm: {e}")

    def _set_product_form_enabled(self, enabled):
        # Khóa/mở các ô nhập của tab Sản phẩm
        for name in ("txtTenSP", "txtGiaBan", "spinSoLuongTon", "cbTrangThai",
                     "txtImagePath", "btnSelectImage"):
            w = getattr(self.ui_product, name, None)
            if w is not None:
                w.setEnabled(enabled)

    def _set_product_edit_button(self, saving):
        # saving=True -> nút hiển thị "Lưu"; ngược lại "Sửa"
        self.ui_product.btnSua.setText("Lưu" if saving else "Sửa")
        if HAS_QTA:
            self.ui_product.btnSua.setIcon(
                qta.icon('fa5s.save' if saving else 'fa5s.edit', color='white'))

    def _product_reset_mode(self):
        self._product_action = None
        self._set_product_form_enabled(False)
        self._set_product_edit_button(False)

    def _on_product_edit_button(self):
        # Nút Sửa/Lưu gộp cho tab Sản phẩm
        if self._product_action == "THEM":
            self._product_do_insert()
        elif self._product_action == "SUA":
            self.product_action_save()
        else:
            if not getattr(self, 'selected_product_id', None):
                QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn một sản phẩm trên bảng để sửa!")
                return
            self._product_action = "SUA"
            self._set_product_form_enabled(True)
            self._set_product_edit_button(True)

    def product_cell_clicked(self, row, column):
        item = self.ui_product.tblProduct.item(row, 0)
        if not item:
            return
        ma_sp = item.text()
        self.selected_product_id = ma_sp

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT TenSP, GiaBan, SoLuongTon, HinhAnh FROM SAN_PHAM WHERE MaSP = ?", (ma_sp,))
            res = cursor.fetchone()
            conn.close()

            if res:
                ten_sp, gia_ban, so_luong, hinh_anh = res
                # Block signals temporarily to prevent infinite recursion during syncing
                self.ui_product.spinSoLuongTon.blockSignals(True)
                self.ui_product.cbTrangThai.blockSignals(True)
                
                self.ui_product.txtTenSP.setText(str(ten_sp or ""))
                self.ui_product.txtGiaBan.setText(str(int(gia_ban) if gia_ban is not None else ""))
                self.ui_product.spinSoLuongTon.setValue(int(so_luong) if so_luong is not None else 0)
                self.ui_product.txtImagePath.setText(str(hinh_anh or ""))
                if so_luong is not None and so_luong > 0:
                    self.ui_product.cbTrangThai.setCurrentIndex(0)
                else:
                    self.ui_product.cbTrangThai.setCurrentIndex(1)

                self.ui_product.spinSoLuongTon.blockSignals(False)
                self.ui_product.cbTrangThai.blockSignals(False)

                self.load_product_image(hinh_anh)
        except Exception as e:
            print(f"Error loading product details: {e}")

    def product_status_changed(self, index):
        self.ui_product.spinSoLuongTon.blockSignals(True)
        if index == 1:  # Hết hàng
            self.ui_product.spinSoLuongTon.setValue(0)
        elif index == 0 and self.ui_product.spinSoLuongTon.value() == 0:
            self.ui_product.spinSoLuongTon.setValue(1)
        self.ui_product.spinSoLuongTon.blockSignals(False)

    def product_spin_changed(self, value):
        self.ui_product.cbTrangThai.blockSignals(True)
        if value == 0:
            self.ui_product.cbTrangThai.setCurrentIndex(1)
        else:
            self.ui_product.cbTrangThai.setCurrentIndex(0)
        self.ui_product.cbTrangThai.blockSignals(False)

    def generate_new_product_id(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT MaSP FROM SAN_PHAM")
            ids = [r[0] for r in cursor.fetchall() if r[0] and r[0].startswith("SP")]
            conn.close()
            max_val = 0
            for i in ids:
                num = i[2:]
                if num.isdigit():
                    max_val = max(max_val, int(num))
            return f"SP{max_val + 1:02d}"
        except Exception as e:
            print(e)
            return "SP99"

    def product_action_them(self):
        # Mở form ở chế độ THÊM: xóa trắng, bật ô nhập, nút Sửa -> Lưu
        self.selected_product_id = None
        self.product_clear_fields()
        self._product_action = "THEM"
        self._set_product_form_enabled(True)
        self._set_product_edit_button(True)

    def _product_do_insert(self):
        ten_sp = self.ui_product.txtTenSP.text().strip()
        gia_ban_str = self.ui_product.txtGiaBan.text().strip()
        so_luong = self.ui_product.spinSoLuongTon.value()
        hinh_anh = self.ui_product.txtImagePath.text().strip()

        if not ten_sp:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Vui lòng nhập Tên sản phẩm!")
            return

        try:
            gia_ban = float(gia_ban_str) if gia_ban_str else 0.0
        except ValueError:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Giá bán phải là số hợp lệ!")
            return

        ma_sp = self.generate_new_product_id()

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("INSERT INTO SAN_PHAM (MaSP, TenSP, GiaBan, SoLuongTon, HinhAnh, MoTa) VALUES (?, ?, ?, ?, ?, ?)",
                           (ma_sp, ten_sp, gia_ban, so_luong, hinh_anh, ""))
            conn.commit()
            conn.close()

            self.load_product_data()
            self.product_clear_fields()
            self._product_reset_mode()
            QMessageBox.information(self, "Thành công", f"Đã thêm sản phẩm {ma_sp} thành công.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể thêm sản phẩm: {e}")

    def product_action_save(self):
        if not hasattr(self, 'selected_product_id') or not self.selected_product_id:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn một sản phẩm trên bảng để sửa!")
            return

        ten_sp = self.ui_product.txtTenSP.text().strip()
        gia_ban_str = self.ui_product.txtGiaBan.text().strip()
        so_luong = self.ui_product.spinSoLuongTon.value()
        hinh_anh = self.ui_product.txtImagePath.text().strip()

        if not ten_sp:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Tên sản phẩm không được để trống!")
            return

        try:
            gia_ban = float(gia_ban_str) if gia_ban_str else 0.0
        except ValueError:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Giá bán phải là số hợp lệ!")
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("UPDATE SAN_PHAM SET TenSP = ?, GiaBan = ?, SoLuongTon = ?, HinhAnh = ? WHERE MaSP = ?",
                           (ten_sp, gia_ban, so_luong, hinh_anh, self.selected_product_id))
            conn.commit()
            conn.close()

            self.load_product_data()
            self._product_reset_mode()
            QMessageBox.information(self, "Thành công", f"Cập nhật thành công sản phẩm {self.selected_product_id}.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể cập nhật sản phẩm: {e}")

    def product_action_xoa(self):
        if not hasattr(self, 'selected_product_id') or not self.selected_product_id:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn một sản phẩm trên bảng để xóa!")
            return

        confirm = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Bạn có chắc chắn muốn xóa sản phẩm {self.selected_product_id} không?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if confirm == QMessageBox.StandardButton.Yes:
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM SAN_PHAM WHERE MaSP = ?", (self.selected_product_id,))
                conn.commit()
                conn.close()

                self.load_product_data()
                self.product_clear_fields()
                self.selected_product_id = None
                QMessageBox.information(self, "Thành công", "Đã xóa sản phẩm khỏi cơ sở dữ liệu.")
            except Exception as e:
                QMessageBox.critical(self, "Lỗi", f"Không thể xóa sản phẩm: {e}")

    def product_export_excel(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Lưu file Excel sản phẩm", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "SanPham"

            headers = ["Mã SP", "Tên Sản Phẩm", "Giá Bán", "Số Tồn Kho", "Đường Dẫn Ảnh", "Trạng Thái"]
            ws.append(headers)

            for row in range(self.ui_product.tblProduct.rowCount()):
                row_data = []
                for col in range(self.ui_product.tblProduct.columnCount()):
                    item = self.ui_product.tblProduct.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            wb.save(path)
            QMessageBox.information(self, "Thành công", f"Đã xuất dữ liệu Excel thành công tại:\n{path}")
        except ImportError:
            QMessageBox.critical(self, "Lỗi", "Vui lòng cài đặt openpyxl bằng lệnh: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể lưu file Excel: {e}")

    def product_action_clear(self):
        self.ui_product.tblProduct.setRowCount(0)
        self.product_clear_fields()
        self.selected_product_id = None
        self._product_reset_mode()
        self.ui_product.lblStatus.setText("Trạng thái: Đã xóa dữ liệu trên bảng hiển thị.")

    def product_clear_fields(self):
        self.ui_product.txtTenSP.clear()
        self.ui_product.txtGiaBan.clear()
        self.ui_product.spinSoLuongTon.setValue(0)
        self.ui_product.txtImagePath.clear()
        self.ui_product.cbTrangThai.setCurrentIndex(0)
        self.ui_product.lblImagePreview.setText("Xem trước hình ảnh")
        self.ui_product.lblImagePreview.setPixmap(QtGui.QPixmap())

    def product_search(self):
        search_text = self.ui_product.txtSearch.text().strip()
        if not search_text:
            self.load_product_data()
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            query = """
                SELECT MaSP, TenSP, GiaBan, SoLuongTon, HinhAnh, 
                       CASE WHEN SoLuongTon > 0 THEN 'Còn hàng' ELSE 'Hết hàng' END 
                FROM SAN_PHAM 
                WHERE TenSP LIKE ?
            """
            cursor.execute(query, (f"%{search_text}%",))
            results = cursor.fetchall()
            conn.close()

            self.populate_table(self.ui_product.tblProduct, results)
            self.ui_product.lblStatus.setText(f"Trạng thái: Tìm thấy {len(results)} sản phẩm.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể tìm kiếm sản phẩm: {e}")

    def product_refresh(self):
        self.ui_product.txtSearch.clear()
        self.load_product_data()
        self.ui_product.lblStatus.setText("Trạng thái: Đã làm mới dữ liệu sản phẩm.")

    def product_select_image(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Chọn hình ảnh sản phẩm", "", "Images (*.png *.jpg *.jpeg *.bmp *.gif)"
        )
        if path:
            filename = os.path.basename(path)
            self.ui_product.txtImagePath.setText(filename)

            try:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                dest = os.path.join(current_dir, filename)
                if not os.path.exists(dest):
                    import shutil
                    shutil.copy(path, dest)
            except Exception as e:
                print(f"Error copying image: {e}")

            self.load_product_image(path)

    # --- PHÂN HỆ NGƯỜI BÁN ---
    def load_seller_data(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT MaNguoiBan, HoTen, SoDienThoai, Email, TenCuaHang FROM NGUOI_BAN")
            self.populate_table(self.ui_seller.tblSeller, cursor.fetchall())
            conn.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể tải dữ liệu người bán: {e}")

    def _set_seller_form_enabled(self, enabled):
        for name in ("txtTenNguoiBan", "txtSoDienThoai", "txtEmail", "txtMatKhau", "txtTenShop"):
            w = getattr(self.ui_seller, name, None)
            if w is not None:
                w.setEnabled(enabled)

    def _set_seller_edit_button(self, saving):
        self.ui_seller.btnSua.setText("Lưu" if saving else "Sửa")
        if HAS_QTA:
            self.ui_seller.btnSua.setIcon(
                qta.icon('fa5s.save' if saving else 'fa5s.edit', color='white'))

    def _seller_reset_mode(self):
        self._seller_action = None
        self._set_seller_form_enabled(False)
        self._set_seller_edit_button(False)

    def _on_seller_edit_button(self):
        if self._seller_action == "THEM":
            self._seller_do_insert()
        elif self._seller_action == "SUA":
            self.seller_action_save()
        else:
            if not getattr(self, 'selected_seller_id', None):
                QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn một người bán trên bảng để sửa!")
                return
            self._seller_action = "SUA"
            self._set_seller_form_enabled(True)
            self._set_seller_edit_button(True)

    def seller_cell_clicked(self, row, column):
        item = self.ui_seller.tblSeller.item(row, 0)
        if not item:
            return
        ma_nb = item.text()
        self.selected_seller_id = ma_nb

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT HoTen, SoDienThoai, Email, MatKhau, TenCuaHang FROM NGUOI_BAN WHERE MaNguoiBan = ?", (ma_nb,))
            res = cursor.fetchone()
            conn.close()

            if res:
                ho_ten, sdt, email, mat_khau, shop = res
                self.ui_seller.txtTenNguoiBan.setText(str(ho_ten or ""))
                self.ui_seller.txtSoDienThoai.setText(str(sdt or ""))
                self.ui_seller.txtEmail.setText(str(email or ""))
                self.ui_seller.txtMatKhau.setText(str(mat_khau or ""))
                self.ui_seller.txtTenShop.setText(str(shop or ""))
        except Exception as e:
            print(f"Error loading seller details: {e}")

    def generate_new_seller_id(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT MaNguoiBan FROM NGUOI_BAN")
            ids = [r[0] for r in cursor.fetchall() if r[0] and r[0].startswith("NB")]
            conn.close()
            max_val = 0
            for i in ids:
                num = i[2:]
                if num.isdigit():
                    max_val = max(max_val, int(num))
            return f"NB{max_val + 1:02d}"
        except Exception as e:
            print(e)
            return "NB99"

    def seller_action_them(self):
        # Mở form ở chế độ THÊM
        self.selected_seller_id = None
        self.seller_clear_fields()
        self._seller_action = "THEM"
        self._set_seller_form_enabled(True)
        self._set_seller_edit_button(True)

    def _seller_do_insert(self):
        ho_ten = self.ui_seller.txtTenNguoiBan.text().strip()
        sdt = self.ui_seller.txtSoDienThoai.text().strip()
        email = self.ui_seller.txtEmail.text().strip()
        mat_khau = self.ui_seller.txtMatKhau.text().strip()
        shop = self.ui_seller.txtTenShop.text().strip()

        if not ho_ten:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Vui lòng nhập Tên người bán!")
            return
        if not mat_khau:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Vui lòng nhập Mật khẩu đăng nhập!")
            return

        ma_nb = self.generate_new_seller_id()

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("INSERT INTO NGUOI_BAN (MaNguoiBan, MatKhau, HoTen, SoDienThoai, Email, TenCuaHang, DiaChi) VALUES (?, ?, ?, ?, ?, ?, ?)",
                           (ma_nb, mat_khau, ho_ten, sdt, email, shop, ""))
            conn.commit()
            conn.close()

            self.load_seller_data()
            self.seller_clear_fields()
            self._seller_reset_mode()
            QMessageBox.information(self, "Thành công", f"Đã thêm người bán {ma_nb} thành công.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể thêm người bán: {e}")

    def seller_action_save(self):
        if not hasattr(self, 'selected_seller_id') or not self.selected_seller_id:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn một người bán trên bảng để sửa!")
            return

        ho_ten = self.ui_seller.txtTenNguoiBan.text().strip()
        sdt = self.ui_seller.txtSoDienThoai.text().strip()
        email = self.ui_seller.txtEmail.text().strip()
        mat_khau = self.ui_seller.txtMatKhau.text().strip()
        shop = self.ui_seller.txtTenShop.text().strip()

        if not ho_ten:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Tên người bán không được để trống!")
            return
        if not mat_khau:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Mật khẩu không được để trống!")
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("UPDATE NGUOI_BAN SET HoTen = ?, SoDienThoai = ?, Email = ?, MatKhau = ?, TenCuaHang = ? WHERE MaNguoiBan = ?",
                           (ho_ten, sdt, email, mat_khau, shop, self.selected_seller_id))
            conn.commit()
            conn.close()

            self.load_seller_data()
            self._seller_reset_mode()
            QMessageBox.information(self, "Thành công", f"Cập nhật thành công người bán {self.selected_seller_id}.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể cập nhật người bán: {e}")

    def seller_action_xoa(self):
        if not hasattr(self, 'selected_seller_id') or not self.selected_seller_id:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn một người bán trên bảng để xóa!")
            return

        confirm = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Bạn có chắc chắn muốn xóa người bán {self.selected_seller_id} không?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if confirm == QMessageBox.StandardButton.Yes:
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM NGUOI_BAN WHERE MaNguoiBan = ?", (self.selected_seller_id,))
                conn.commit()
                conn.close()

                self.load_seller_data()
                self.seller_clear_fields()
                self.selected_seller_id = None
                QMessageBox.information(self, "Thành công", "Đã xóa người bán khỏi cơ sở dữ liệu.")
            except Exception as e:
                QMessageBox.critical(self, "Lỗi", f"Không thể xóa người bán: {e}")

    def seller_export_excel(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Lưu file Excel người bán", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "NguoiBan"

            headers = ["Mã Người Bán", "Tên Người Bán", "Số Điện Thoại", "Email", "Tên Shop"]
            ws.append(headers)

            for row in range(self.ui_seller.tblSeller.rowCount()):
                row_data = []
                for col in range(self.ui_seller.tblSeller.columnCount()):
                    item = self.ui_seller.tblSeller.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            wb.save(path)
            QMessageBox.information(self, "Thành công", f"Đã xuất dữ liệu Excel thành công tại:\n{path}")
        except ImportError:
            QMessageBox.critical(self, "Lỗi", "Vui lòng cài đặt openpyxl bằng lệnh: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể lưu file Excel: {e}")

    def seller_action_clear(self):
        self.ui_seller.tblSeller.setRowCount(0)
        self.seller_clear_fields()
        self.selected_seller_id = None
        self._seller_reset_mode()
        self.ui_seller.lblStatus.setText("Trạng thái: Đã xóa dữ liệu trên bảng hiển thị.")

    def seller_clear_fields(self):
        self.ui_seller.txtTenNguoiBan.clear()
        self.ui_seller.txtSoDienThoai.clear()
        self.ui_seller.txtEmail.clear()
        self.ui_seller.txtMatKhau.clear()
        self.ui_seller.txtTenShop.clear()

    def seller_search(self):
        search_text = self.ui_seller.txtSearch.text().strip()
        if not search_text:
            self.load_seller_data()
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            query = """
                SELECT MaNguoiBan, HoTen, SoDienThoai, Email, TenCuaHang 
                FROM NGUOI_BAN 
                WHERE HoTen LIKE ? OR TenCuaHang LIKE ?
            """
            cursor.execute(query, (f"%{search_text}%", f"%{search_text}%"))
            results = cursor.fetchall()
            conn.close()

            self.populate_table(self.ui_seller.tblSeller, results)
            self.ui_seller.lblStatus.setText(f"Trạng thái: Tìm thấy {len(results)} người bán.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể tìm kiếm người bán: {e}")

    def seller_refresh(self):
        self.ui_seller.txtSearch.clear()
        self.load_seller_data()
        self.ui_seller.lblStatus.setText("Trạng thái: Đã làm mới dữ liệu người bán.")

    def search_table_generic(self, table_widget, db_table, pk_col, search_text, refresh_callback):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(f"PRAGMA table_info({db_table})")
            cols = [row[1] for row in cursor.fetchall()]
            if not cols:
                conn.close()
                return

            conditions = []
            params = []
            for col in cols:
                conditions.append(f"{col} LIKE ?")
                params.append(f"%{search_text}%")

            query = f"SELECT * FROM {db_table} WHERE {' OR '.join(conditions)}"
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()

            self.populate_table(table_widget, results)
        except Exception as e:
            print(f"Error generic searching: {e}")


class LoginEx(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_LoginDialog()
        self.ui.setupUi(self)

        self.center_window()
        self.setup_signals()
        apply_fontawesome_icons(self)
        # Logo app trên thẻ đăng nhập (nền gradient tròn) — biểu tượng phát trực tiếp
        if HAS_QTA and hasattr(self.ui, "lblLogo"):
            self.ui.lblLogo.setText("")
            self.ui.lblLogo.setPixmap(qta.icon("fa5s.broadcast-tower", color="#ffffff").pixmap(QtCore.QSize(38, 38)))
            self.ui.lblLogo.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

    def center_window(self):
        qr = self.frameGeometry()
        cp = QtGui.QGuiApplication.primaryScreen().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def setup_signals(self):
        self.ui.btnLogin.clicked.connect(self.process_login)
        self.ui.btnExit.clicked.connect(self.close)

    def process_login(self):
        username = self.ui.txtUsername.text().strip()
        password = self.ui.txtPassword.text().strip()

        if not username or not password:
            self.ui.lblStatus.setText("Vui lòng điền đủ tài khoản và mật khẩu!")
            return

        try:
            # Kết nối Database SQLite kiểm tra thông tin đăng nhập thực tế
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # Truy vấn khớp thông tin từ bảng NGUOI_BAN
            cursor.execute("SELECT MaNguoiBan, HoTen FROM NGUOI_BAN WHERE MaNguoiBan=? AND MatKhau=?",
                           (username, password))
            result = cursor.fetchone()
            conn.close()

            if result:
                # Đăng nhập thành công, truyền thông tin sang Dashboard
                seller_id, seller_name = result[0], result[1]
                self.open_dashboard_window(seller_id, seller_name)
            else:
                self.ui.lblStatus.setText("Sai mã người bán hoặc mật khẩu!")

        except sqlite3.OperationalError as e:
            # Bypass trường hợp cấu trúc bảng thay đổi hoặc không tìm thấy cơ sở dữ liệu
            self.ui.lblStatus.setText(f"Lỗi DB: {e}. Thử admin/123456")
            if username == "NB01" or username == "admin":
                self.open_dashboard_window(username, "Người Bán Bypass")

    def open_dashboard_window(self, seller_id, seller_name):
        self.dashboard_window = DashboardEx(seller_id, seller_name, login_window=self)
        self.dashboard_window.show()
        self.hide()
class PaymentEx(QtWidgets.QWidget, Ui_PaymentWidget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        # --- CẤU HÌNH BẢNG TỰ ĐỘNG BẮT LỰA CHỌN DÒNG ---
        self.tblPayment.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.tblPayment.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        self.tblPayment.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)

        # Dữ liệu giả lập danh mục Đơn hàng nạp vào ComboBox
        self.list_orders = ["DH001", "DH002", "DH003", "DH004"]
        self.cbDonHang.addItems(self.list_orders)

        # Dữ liệu giả lập danh sách Thanh toán ban đầu
        self.data_payments = [
            {"pay_id": "GD001", "order_id": "DH001", "date": "2026-07-15 08:30:00", "amount": 470000,
             "method": "Chuyển khoản", "status": "Đã thanh toán"},
            {"pay_id": "GD002", "order_id": "DH002", "date": "2026-07-15 09:15:00", "amount": 250000,
             "method": "Tiền mặt", "status": "Đã thanh toán"},
            {"pay_id": "GD003", "order_id": "DH003", "date": "2026-07-15 10:00:00", "amount": 1200000,
             "method": "Ví điện tử", "status": "Chưa thanh toán"}
        ]

        # Biến quản lý trạng thái thao tác: None, "THEM", hoặc "SUA"
        self.current_action = None

        # Kết nối sự kiện tương tác của các nút bấm chức năng
        self.btnSearch.clicked.connect(self.action_search)
        self.btnRefresh.clicked.connect(self.action_refresh)
        self.btnThem.clicked.connect(self.action_them)
        self.btnSua.clicked.connect(self.action_sua)
        self.btnXoa.clicked.connect(self.action_xoa)
        self.btnLuu.clicked.connect(self.action_luu)
        self.btnHuy.clicked.connect(self.action_huy)
        self.btnXuatExcel.clicked.connect(self.action_xuat_excel)

        # Sự kiện click chọn dòng trên TableWidget tự động nhảy thông tin lên Form nhập liệu
        self.tblPayment.itemSelectionChanged.connect(self.sync_table_to_form)

        # Đặt thời gian mặc định cho ô QDateTimeEdit là thời gian hiện tại
        self.dateThanhToan.setDateTime(QtCore.QDateTime.currentDateTime())

        # Tải dữ liệu mặc định hiển thị lên bảng và tạm khóa Form nhập liệu ban đầu
        self.load_data_to_table(self.data_payments)
        self.set_form_enabled(False)

    def load_data_to_table(self, data_list):
        """Hiển thị danh sách giao dịch thanh toán lên QTableWidget"""
        self.tblPayment.setRowCount(0)
        for row_idx, row_data in enumerate(data_list):
            self.tblPayment.insertRow(row_idx)
            self.tblPayment.setItem(row_idx, 0, QTableWidgetItem(row_data["pay_id"]))
            self.tblPayment.setItem(row_idx, 1, QTableWidgetItem(row_data["order_id"]))
            self.tblPayment.setItem(row_idx, 2, QTableWidgetItem(row_data["date"]))
            self.tblPayment.setItem(row_idx, 3, QTableWidgetItem(f"{row_data['amount']:,}"))
            self.tblPayment.setItem(row_idx, 4, QTableWidgetItem(row_data["method"]))
            self.tblPayment.setItem(row_idx, 5, QTableWidgetItem(row_data["status"]))

        # Tự co giãn các cột dữ liệu theo kích thước hiển thị của màn hình
        self.tblPayment.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)

    def set_form_enabled(self, enabled=True):
        """Đóng/Mở trạng thái chỉnh sửa các ô nhập liệu tùy thuộc hành động"""
        self.cbDonHang.setEnabled(enabled)
        self.dateThanhToan.setEnabled(enabled)
        self.txtSoTien.setEnabled(enabled)
        self.cbPhuongThuc.setEnabled(enabled)
        self.cbTrangThai.setEnabled(enabled)

        # Khóa/Mở các nút chức năng tương ứng để ngăn ngừa xung đột dữ liệu
        self.btnThem.setEnabled(not enabled)
        self.btnSua.setEnabled(not enabled)
        self.btnXoa.setEnabled(not enabled)
        self.btnLuu.setEnabled(enabled)
        self.btnHuy.setEnabled(enabled)

    def sync_table_to_form(self):
        """Khi click chọn dòng trên bảng, đổ thông tin tương ứng sang khung bên cạnh"""
        if self.current_action is not None:
            return  # Đang trong trạng thái Thêm/Sửa thì không tự động chuyển dòng khi click bảng

        selected_rows = self.tblPayment.selectionModel().selectedRows()
        if not selected_rows:
            return

        row = selected_rows[0].row()
        order_id = self.tblPayment.item(row, 1).text()
        date_str = self.tblPayment.item(row, 2).text()
        amount = self.tblPayment.item(row, 3).text().replace(",", "")
        method = self.tblPayment.item(row, 4).text()
        status = self.tblPayment.item(row, 5).text()

        # Chọn item tương ứng trên Form
        self.cbDonHang.setCurrentText(order_id)

        q_datetime = QtCore.QDateTime.fromString(date_str, "yyyy-MM-dd HH:mm:ss")
        if q_datetime.isValid():
            self.dateThanhToan.setDateTime(q_datetime)

        self.txtSoTien.setText(amount)
        self.cbPhuongThuc.setCurrentText(method)
        self.cbTrangThai.setCurrentText(status)

    def clear_form(self):
        """Xóa dữ liệu cũ trên form về mặc định để chuẩn bị thao tác mới"""
        if self.cbDonHang.count() > 0: self.cbDonHang.setCurrentIndex(0)
        self.dateThanhToan.setDateTime(QtCore.QDateTime.currentDateTime())
        self.txtSoTien.clear()
        self.cbPhuongThuc.setCurrentIndex(0)
        self.cbTrangThai.setCurrentIndex(0)

    # --- ĐỊNH NGHĨA CÁC HÀM XỬ LÝ SỰ KIỆN NÚT BẤM ---

    def action_search(self):
        """Tìm kiếm giao dịch theo Mã Đơn Hàng hoặc Phương Thức"""
        search_text = self.txtSearch.text().strip().lower()
        if not search_text:
            self.load_data_to_table(self.data_payments)
            return

        filtered_data = [
            item for item in self.data_payments
            if search_text in item["order_id"].lower() or search_text in item["method"].lower()
        ]
        self.load_data_to_table(filtered_data)
        self.lblStatus.setText(f"Trạng thái: Đã tìm thấy {len(filtered_data)} kết quả giao dịch.")

    def action_refresh(self):
        """Làm mới ô tìm kiếm và đặt danh sách bảng về trạng thái mặc định"""
        self.txtSearch.clear()
        self.load_data_to_table(self.data_payments)
        self.lblStatus.setText("Trạng thái: Đã làm mới danh sách thanh toán.")

    def action_them(self):
        """Kích hoạt trạng thái thêm mới giao dịch thanh toán"""
        self.current_action = "THEM"
        self.clear_form()
        self.set_form_enabled(True)
        self.txtSoTien.setFocus()
        self.lblStatus.setText("Trạng thái: Đang nhập mới thông tin giao dịch...")

    def action_sua(self):
        """Kích hoạt trạng thái sửa đổi thông tin giao dịch đang chọn"""
        selected_rows = self.tblPayment.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn một giao dịch trên bảng để sửa!")
            return

        self.current_action = "SUA"
        self.set_form_enabled(True)
        self.lblStatus.setText("Trạng thái: Đang chỉnh sửa thông tin giao dịch thanh toán...")

    def action_xoa(self):
        """Xóa bỏ bản ghi giao dịch thanh toán đang được lựa chọn"""
        selected_rows = self.tblPayment.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn một giao dịch trên bảng để xóa!")
            return

        row = selected_rows[0].row()
        pay_id = self.tblPayment.item(row, 0).text()

        confirm = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Bạn có chắc chắn muốn xóa mã giao dịch {pay_id} không?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if confirm == QMessageBox.StandardButton.Yes:
            self.data_payments = [item for item in self.data_payments if item["pay_id"] != pay_id]
            self.load_data_to_table(self.data_payments)
            self.clear_form()
            self.lblStatus.setText(f"Trạng thái: Đã xóa thành công mã giao dịch {pay_id}")

    def action_luu(self):
        """Lưu trữ thông tin dữ liệu sau khi nhấn Thêm mới hoặc Chỉnh sửa"""
        order_id = self.cbDonHang.currentText()
        date_str = self.dateThanhToan.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        method = self.cbPhuongThuc.currentText()
        status = self.cbTrangThai.currentText()

        try:
            amount_text = self.txtSoTien.text().replace(",", "").strip()
            amount = int(amount_text) if amount_text else 0
        except ValueError:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Số tiền thanh toán phải là số nguyên hợp lệ!")
            return

        if amount <= 0:
            QMessageBox.warning(self, "Lỗi dữ liệu", "Số tiền thanh toán phải lớn hơn 0!")
            return

        if self.current_action == "THEM":
            # Tự động sinh mã giao dịch tăng tiến tiếp theo
            new_id = f"GD{len(self.data_payments) + 1:03d}"
            new_item = {
                "pay_id": new_id,
                "order_id": order_id,
                "date": date_str,
                "amount": amount,
                "method": method,
                "status": status
            }
            self.data_payments.append(new_item)
            self.lblStatus.setText(f"Trạng thái: Thêm mới thành công giao dịch {new_id}")

        elif self.current_action == "SUA":
            selected_rows = self.tblPayment.selectionModel().selectedRows()
            row = selected_rows[0].row()
            pay_id = self.tblPayment.item(row, 0).text()

            # Tiến hành cập nhật thông tin chỉnh sửa vào mảng dữ liệu
            for item in self.data_payments:
                if item["pay_id"] == pay_id:
                    item["order_id"] = order_id
                    item["date"] = date_str
                    item["amount"] = amount
                    item["method"] = method
                    item["status"] = status
                    break
            self.lblStatus.setText(f"Trạng thái: Đã cập nhật thành công giao dịch {pay_id}")

        self.load_data_to_table(self.data_payments)
        self.current_action = None
        self.set_form_enabled(False)

    def action_huy(self):
        """Hủy bỏ thao tác hiện tại"""
        self.current_action = None
        self.clear_form()
        self.set_form_enabled(False)
        self.sync_table_to_form()
        self.lblStatus.setText("Trạng thái: Đã hủy bỏ thao tác nghiệp vụ.")

    def action_xuat_excel(self):
        """Kết xuất cấu trúc dữ liệu của bảng thanh toán ra file CSV tiện ích"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Lưu file dữ liệu báo cáo", "", "CSV Files (*.csv);;All Files (*)"
        )
        if path:
            try:
                with open(path, 'w', encoding='utf-8-sig') as f:
                    # Ghi dòng tiêu đề cột dữ liệu
                    headers = ["Mã Giao Dịch", "Mã Đơn Hàng", "Ngày Thanh Toán", "Số Tiền (VND)", "Phương Thức",
                               "Trạng Thái"]
                    f.write(",".join(headers) + "\n")
                    # Ghi nội dung chi tiết của từng bản ghi dữ liệu
                    for item in self.data_payments:
                        row_str = f"{item['pay_id']},{item['order_id']},{item['date']},{item['amount']},{item['method']},{item['status']}"
                        f.write(row_str + "\n")
                QMessageBox.information(self, "Xuất dữ liệu", f"Đã kết xuất dữ liệu thành công tại đường dẫn:\n{path}")
                self.lblStatus.setText("Trạng thái: Xuất file báo cáo thành công.")
            except Exception as e:
                QMessageBox.critical(self, "Lỗi hệ thống", f"Không thể lưu trữ tệp tin: {str(e)}")

def force_light_theme(app):
    """UI thiết kế nền sáng nhưng macOS dark mode đổi chữ mặc định thành trắng — ép theme sáng."""
    app.setStyle("Fusion")
    palette = QtGui.QPalette()
    palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor("#f8fafc"))
    palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtGui.QColor("#000000"))
    palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor("#ffffff"))
    palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor("#f1f5f9"))
    palette.setColor(QtGui.QPalette.ColorRole.Text, QtGui.QColor("#000000"))
    palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor("#e2e8f0"))
    palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtGui.QColor("#000000"))
    palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtGui.QColor("#ffffff"))
    palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtGui.QColor("#000000"))
    palette.setColor(QtGui.QPalette.ColorRole.PlaceholderText, QtGui.QColor("#94a3b8"))
    palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor("#7c3aed"))
    palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtGui.QColor("#ffffff"))
    app.setPalette(palette)


def main():
    app = QApplication(sys.argv)
    force_light_theme(app)
    login_window = LoginEx()
    login_window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()